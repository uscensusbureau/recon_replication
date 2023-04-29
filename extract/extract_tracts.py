'''Extract selected tracts from rHDF and compute uniqueness within rHDF'''

import os
import pandas as pd
import openpyxl as xl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.comments import Comment
from openpyxl.styles import Font
from stats import Stats

AUTH=''

BINMAP = {0: '0', 1: '1', 2: '2', 3: '3', 4: '4', 5: '5', 6: '6', 7: '7',
          8: '8', 9: '9', 10: '10', 11: '11', 12: '12', 13: '13', 14: '14',
          15: '15', 16: '16', 17: '17', 18: '18', 19: '19', 20: '20', 21: '21',
          22: '22 - 24', 25: '25 - 29', 30: '30 - 34', 35: '35 - 39',
          40: '40 - 44', 45: '45 - 49', 50: '50 - 54', 55: '55 - 59',
          60: '60 - 61', 62: '62 - 64', 65: '65 - 66', 67: '67 - 69',
          70: '70 - 74', 75: '75 - 79', 80: '80 - 84', 85: '85+'}

#note, selected have to leave out leading 0
SELECTED = list(set(["42059970400", "17079977500", "46135966400",
                     "31011960100", "31067964700", "31095963600",
                     "01073005103", "48427950105",
                     "34017006900", "06037532400"] +
                    ["38103960000", "38009952400", "19137960100", "19003950100",
                     "31023967800", "20181453600", "27073180200", "04013817400",
                     "20157978100", "08075965900", "38005956500", "04013040525",
                     "29115490100", "04013040514", "38087965000", "19173180300",
                     "04013615500", "27149480100" ]))


#schema
#geoid_tract,geoid_block,sex,age,white,black,aian,asian,nhopi,sor,hisp
COLDEF = {'geoid_tract': 'str',
          'geoid_block': 'str',
          'sex': 'str',
          'age': 'int',
          'white': 'str',
          'black': 'str',
          'aian': 'str',
          'asian': 'str',
          'nhopi': 'str',
          'sor': 'str',
          'hisp': 'str'}

BOLD = Font(bold=True)

def columns_best_fit(ws: xl.worksheet.worksheet.Worksheet):
    """
    Make all columns best fit
    """
    column_letters = tuple(xl.utils.get_column_letter(col_number + 1) for col_number in range(ws.max_column))
    for column_letter in column_letters:
        ws.column_dimensions[column_letter].bestFit = True


def export_xl(df: pd.DataFrame, dfname: str, out_path: str):
    wb = xl.Workbook()
    ws = wb.active
    ws.title = '{} Extract'.format(dfname)
    for r in dataframe_to_rows(df, index=True, header=True):
            ws.append(r)
    if ws['B2'].value is None:
        ws.delete_rows(2)
    ws['A1'] = 'RowID'
    ws['A1'].comment = Comment('Row ID number', AUTH)
    ws['B1'].comment = Comment('Census Block ID number', AUTH)
    ws['C1'].comment = Comment('Sex', AUTH)
    ws['D1'].comment = Comment('Age in years', AUTH)
    ws['E1'].comment = Comment('Binned age', AUTH)
    ws['F1'].comment = Comment('Race contains White', AUTH)
    ws['G1'].comment = Comment('Race contains Black', AUTH)
    ws['H1'].comment = Comment('Race contains American Indian or Alaskan Native', AUTH)
    ws['I1'].comment = Comment('Race contains Asian', AUTH)
    ws['J1'].comment = Comment('Race contains Native Hawaiian or Pacific Islander', AUTH)
    ws['K1'].comment = Comment('Race contains Some Other Race', AUTH)
    ws['L1'].comment = Comment('Hispanic indicator', AUTH)
    ws['M1'].comment = Comment('Record is unique on block, sex, and binned age in {}'.format(dfname), AUTH)
    ws['N1'].comment = Comment('Record is unique on block, sex, binned age, ethnicity (Hispanic), and race in {}'.format(dfname), AUTH)
    for row in ws['A1:N1']:
        for cell in row:
            cell.font = BOLD
    columns_best_fit(ws)
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['N'].width = 14
    ws.freeze_panes = ws['O2']
    wb.save(out_path)

def create_extract(rhdf_path: str):

    rHDF = pd.read_csv(rhdf_path, dtype=COLDEF).query('geoid_tract in @SELECTED')

    rHDF['tract'] = rHDF['geoid_tract']
    rHDF['block'] = rHDF['geoid_block']
    #Put on binned age
    rHDF['binage'] = [Stats.binAge(None, x) for x in rHDF.age.values]

    #Get frequencies for BSAb, BER, and BSAbER
    BSAb = pd.crosstab(rHDF.block, [rHDF.binage, rHDF.sex])
    BER = pd.crosstab(rHDF.block, [rHDF.white, rHDF.black, rHDF.aian,
                                         rHDF.asian, rHDF.nhopi, rHDF.sor,
                                         rHDF.hisp])
    BSAbER = pd.crosstab(rHDF.block, [rHDF.binage, rHDF.sex,
                                            rHDF.white, rHDF.black, rHDF.aian,
                                            rHDF.asian, rHDF.nhopi, rHDF.sor,
                                            rHDF.hisp])

    #Can access counts using:
    #BSAb.loc[<block>].loc[(<binage>, <sex>)
    rHDF['BSAb_count'] = [BSAb.loc[rHDF.block.iloc[idx]].
                          loc[(rHDF.binage.iloc[idx], rHDF.sex.iloc[idx])]
                          for idx in range(0, rHDF.shape[0])]
     
    rHDF['BER_count'] = [BER.loc[rHDF.block.iloc[idx]].
                         loc[(rHDF.white.iloc[idx], rHDF.black.iloc[idx],
                              rHDF.aian.iloc[idx], rHDF.asian.iloc[idx],
                              rHDF.nhopi.iloc[idx], rHDF.sor.iloc[idx],
                              rHDF.hisp.iloc[idx])]
                         for idx in range(0, rHDF.shape[0])]

    rHDF['BSAbER_count'] = [BSAbER.loc[rHDF.block.iloc[idx]].
                            loc[(rHDF.binage.iloc[idx], rHDF.sex.iloc[idx],
                                 rHDF.white.iloc[idx], rHDF.black.iloc[idx],
                                 rHDF.aian.iloc[idx], rHDF.asian.iloc[idx],
                                 rHDF.nhopi.iloc[idx], rHDF.sor.iloc[idx],
                                 rHDF.hisp.iloc[idx])]
                            for idx in range(0, rHDF.shape[0])]

    out = pd.DataFrame()
    out['Block ID'] = rHDF.block
    out['Sex'] = rHDF.sex
    out['Age'] = rHDF.age
    out['Binned Age'] = [BINMAP.get(x) for x in rHDF.binage.values]
    out['White'] = rHDF.white
    out['Black'] = rHDF.black
    out['American Indian or Alaskan Native'] = rHDF.aian
    out['Asian'] = rHDF.asian
    out['Native Hawaiian or Pacific Island'] = rHDF.nhopi
    out['Some Other Race'] = rHDF.sor
    out['Hispanic'] = rHDF.hisp
    out['BSAb Unique'] = rHDF.BSAb_count.eq(1)
    out['BSAbER Unique'] = rHDF.BSAbER_count.eq(1)

    out.reset_index(inplace=True)
    out.drop(columns = 'index', inplace=True)
    out_dir = os.path.dirname(rhdf_path)
    in_file, in_ext = os.path.splitext(os.path.basename(rhdf_path))
    out_name = "{}_0solvar_extract{}".format(in_file, in_ext)
    out_path = os.path.join(out_dir, out_name)
#    out.to_csv(out_path, index=True)
    xl_name = "{}_0solvar_extract.xlsx".format(in_file)
    xl_path = os.path.join(out_dir, xl_name)
    export_xl(out, in_file, xl_path)

if __name__=="__main__":
        import argparse

        parser = argparse.ArgumentParser(description='Extracts a list of tracts containing only blocks with zero solution variability',
                                         formatter_class=argparse.ArgumentDefaultsHelpFormatter)
        parser.add_argument("rhdf_path", type=str, help="rHDF file from which to extract tracts")
        args = parser.parse_args()
        create_extract(args.rhdf_path)

