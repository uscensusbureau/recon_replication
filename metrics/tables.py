#!/usr/bin/env python
# coding: utf-8

import argparse
parser = argparse.ArgumentParser()
parser.add_argument("-v","--version", help="pick version of run", required=True)
parser.add_argument("-r", "--unitfileready", help = "is unit file ready?", required=True)
parser.add_argument("-u","--unitversion", help ="pick unit version of run", required = False, default='False')
parser.add_argument("-p","--prfileready", help ="is PR included in person metrics?", required = False, default='False')
parser.add_argument("-s","--relfile", help ="use separate RELSHIP metrics file?", required = False, default='False')
args = parser.parse_args()

usecases_ready = True
if args.unitfileready == "True":
    unitmetricsready = True
    prunitmetricsready = True
else:
    unitmetricsready = False
    prunitmetricsready = False
if args.prfileready == "False":
    prmetricsready = False
    prunitmetricsready = False
else:
    prmetricsready = True

if args.unitfileready == "True" and args.prfileready == "False":
    prunitmetricsready = False

tractnumracessizesready = True ## TODO: Fix this in the metrics code for DHC
airready = True

if args.relfile == "True":
    userelfile = True
else:
    userelfile = False

#BASEDIR = "Test_DAS_20230228"
# BASEDIR = "Production_DAS_2020_20221123"
BASEDIR = "./output"
FILETYPE = "hdf"
# BASEDIR = "Test_DAS_20220825-DDP2"
# BASEDIR = "Test_Rolando_20220224"
TABLESHELLDIR = BASEDIR
INPUTFILEEND = ""
OUTPUTDIR = f"{BASEDIR}/{args.version}"
OUTPUTFILEEND = f"_{args.version}"
if args.unitversion == "False":
    UNITDIR = f"{BASEDIR}/{args.version}"
else:
    UNITDIR = f"{BASEDIR}/{args.unitversion}"
# OUTPUTDIR = ""
# OUTPUTFILEEND = ""

import pandas as pd

racealonedict  = {1: 'White Alone',  2: 'Black Alone',  3: 'AIAN Alone',  4: 'Asian Alone',  5: 'NHOPI Alone',  6: 'Some Other Race Alone',  7: 'Two Or More Races'}
hispdict = {'1': 'Not Hispanic', '2':'Hispanic'}
hhtypeownchildlt18dict = {'1': 'Married couple household: With own children < 18', '2': 'Cohabiting couple household: With own children < 18','3': 'Male householder, no spouse/partner present: With own children < 18', '4':'Female householder, no spouse/partner present: With own children < 18'}
tenuredict = {'1': 'Owned with a mortgage', '2':'Owned free and clear', '3':'Rented', '4':'Occupied without payment of rent'}
vacantdict = {'1': 'For rent', '2': 'Rented, not occupied', '3': 'For sale only', '4': 'Sold, not occupied', '5': 'For seasonal, recreational, or occasional use', '6': 'For migrant workers', '7': 'Other vacant' }
cpltdict = {'1': 'Opposite-sex married couple household', '2': 'Same-sex married couple household', '3': 'Opposite-sex unmarried partner household', '4': 'Same-sex unmarried partner household',}
hheragedict = { '1':'Householder 15 to 24 years', '2':'Householder 25 to 34 years', '3':'Householder 35 to 54 years', '4':'Householder 55 to 64 years', '5':'Householder 65 years and over', }
reldict = {20:"Householder", 21: "Opposite-sex spouse", 22: "Opposite-sex unmarried partner", 23: "Same-sex spouse", 24: "Same-sex unmarried partner", 25: "Biological son or daughter", 26: "Adopted son or daughter", 27: "Stepson or stepdaughter", 28: "Brother or sister", 29: "Father or mother", 30: "Grandchild", 31:"Parent-in-law", 32:"Son-in-law or daughter-in-law", 33:"Other relative", 34: "Roommate or housemate", 35: "Foster child", 36: "Other nonrelative", #37: "Institutional GQ person", 38: "Noninstitutional GQ person",
}


from openpyxl import load_workbook
wb = load_workbook(filename = f"{BASEDIR}/table_shells.xlsx")
# wsrm = wb["Readme"]
# if unitmetricsready:
#     readme = pd.read_csv(f"../output/{OUTPUTDIR}/filemissionnames_unitmetrics.csv")
#     wsrm['A1'] = f"Person Mission Name: {readme['PersonMissionName'].iat[0]}" 
#     wsrm['A2'] = f"Unit Mission Name: {readme['UnitMissionName'].iat[0]}"
#     try:
#         wsrm['A3'] = f"Puerto Rico Person Mission Name: {readme['PersonPRMissionName'].iat[0]}"
#     except KeyError:
#         wsrm['A3'] = f"Puerto Rico Person Mission Name: N/A"
#     try:
#         wsrm['A4'] = f"Puerto Rico Unit Mission Name: {readme['UnitPRMissionName'].iat[0]}"
#     except KeyError:
#         wsrm['A4'] = f"Puerto Rico Unit Mission Name: N/A"
# else:
#     readme = pd.read_csv(f"../output/{OUTPUTDIR}/filemissionnames_personmetrics.csv")
#     wsrm['A1'] = f"Person Mission Name: {readme['PersonMissionName'].iat[0]}"
#     try:
#         wsrm['A3'] = f"Puerto Rico Person Mission Name: {readme['PersonPRMissionName'].iat[0]}"
#     except KeyError:
#         wsrm['A3'] = f"Puerto Rico Person Mission Name: N/A"
ws = wb["Updated BAP"]

#df = pd.read_csv(f"../output/{OUTPUTDIR}/{FILETYPE}_per_metrics_v2{INPUTFILEEND}.csv")
#df = pd.read_csv(f"../output/PPMF20210428/hdf_per_metrics_v2.csv")
df = pd.read_csv(f"{OUTPUTDIR}/hdf_per_metrics_v2.csv")
if unitmetricsready:
    dfu = pd.read_csv(f"../output/{UNITDIR}/{FILETYPE}_unit_metrics_v2{INPUTFILEEND}.csv")
else:
    dfu = None
if userelfile:
    dfrel = pd.read_csv(f"../output/{OUTPUTDIR}/{FILETYPE}_relship_metrics_v2{INPUTFILEEND}.csv")
elif unitmetricsready:
    dfrel = dfu.copy()
else:
    dfrel = None


# totalpopsizecats = ['[0-1000)', '[1000 - 5000)', '[5000 - 10000)', '[10000 - 50000)', '[50000 - 100000)', '[100000 - Inf)']
# totalpopsizecats_plusall = ['All','[0-1000)', '[1000 - 5000)', '[5000 - 10000)', '[10000 - 50000)', '[50000 - 100000)', '[100000 - Inf)']
sizecats = ['[0.0, 1000.0)', '[1000.0, 5000.0)', '[5000.0, 10000.0)', '[10000.0, 50000.0)', '[50000.0, 100000.0)', '[100000.0, inf)']
sizecats_plusall = ['All','[0.0, 1000.0)', '[1000.0, 5000.0)', '[5000.0, 10000.0)', '[10000.0, 50000.0)', '[50000.0, 100000.0)', '[100000.0, inf)']
incplacesizecats_plusall = ['All','[0.0, 500.0)','[500.0, 1000.0)', '[1000.0, 5000.0)', '[5000.0, 10000.0)', '[10000.0, 50000.0)', '[50000.0, 100000.0)', '[100000.0, inf)']
if airready:
    airanvsasizecats_plusall = ['All', '[0.0, 100.0)', '[100.0, 1000.0)', '[1000.0, 10000.0)', '[10000.0, inf)']
else:
    airanvsasizecats_plusall = ['All', '[0.0, 100.0)', '[100.0, 1000.0)', '[1000.0, inf)', '[1000.0, inf)']
smallsizecats = ['[0.0, 10.0)', '[10.0, 100.0)', '[100.0, inf)']
smallsizecats_plusall = ['All','[0.0, 10.0)', '[10.0, 100.0)', '[100.0, inf)']
sizecats_u18_plusall = ['All','[0.0, 1000.0)', '[1000.0, 10000.0)', '[10000.0, inf)']
racealonecats = ["White Alone", "Black Alone", "AIAN Alone", "Asian Alone", "NHOPI Alone", "Some Other Race Alone", "Two Or More Races"]
raceincombcats = ["White Alone or In Combination", "Black Alone or In Combination", "AIAN Alone or In Combination", "Asian Alone or In Combination", "NHOPI Alone or In Combination", "Some Other Race Alone or In Combination"]
numracecats = [ '1 Race(s)', '2 Race(s)', '3 Race(s)', '4 Race(s)', '5 Race(s)', '6 Race(s)']


# Table 1a
row = 7
for s in sizecats_plusall:
    ws['B'+str(row)] = df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'NumCells'].iat[0]
    ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'MeanAbsDiff'].iat[0], 2)
    ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'RMSE'].iat[0], 2)
    ws['E'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'MeanAbsPercDiff'].iat[0], 2)
    ws['F'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'CV'].iat[0], 2)
    ws['G'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'MeanPercDiff'].iat[0], 2)
    ws['H'+str(row)] = df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'NumberBtw2Perc5Perc'].iat[0]
    ws['I'+str(row)] = df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'NumberGreater5Perc'].iat[0]
    ws['J'+str(row)] = df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'NumberGreater200'].iat[0]
    row += 1

## Table 1b
row = 19
for s in incplacesizecats_plusall:
    ws['B'+str(row)] = df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'NumCells'].iat[0]
    ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'MeanAbsDiff'].iat[0], 2)
    ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'RMSE'].iat[0], 2)
    ws['E'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'MeanAbsPercDiff'].iat[0], 2)
    ws['F'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'CV'].iat[0], 2)
    ws['G'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'MeanPercDiff'].iat[0], 2)
    ws['H'+str(row)] = df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'NumberBtw2Perc5Perc'].iat[0]
    ws['I'+str(row)] = df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'NumberGreater5Perc'].iat[0]
    ws['J'+str(row)] = df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'NumberGreater200'].iat[0]
    row += 1

# Table 1c
row = 32
ws['B'+str(row)] = df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population"), 'NumCells'].iat[0]
ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population"), 'MeanAbsDiff'].iat[0], 2)
ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population"), 'RMSE'].iat[0], 2)
ws['E'+str(row)] = df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population"), 'NumberBtw2Perc5Perc'].iat[0]
ws['F'+str(row)] = df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population"), 'NumberGreater5Perc'].iat[0]
ws['G'+str(row)] = df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population"), 'NumberGreater200'].iat[0]

# Table 1d
row = 38
ws['B'+str(row)] = df.loc[(df['Geography'] == 'Block') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population for Urban Blocks"), 'NumCells'].iat[0]
ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'Block') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population for Urban Blocks"), 'MeanAbsDiff'].iat[0], 2)
ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'Block') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population for Urban Blocks"), 'RMSE'].iat[0], 2)
ws['E'+str(row)] = round(df.loc[(df['Geography'] == 'Block') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population for Urban Blocks"), 'MeanAbsPercDiff'].iat[0], 2)
ws['F'+str(row)] = round(df.loc[(df['Geography'] == 'Block') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population for Urban Blocks"), 'CV'].iat[0], 2)
ws['G'+str(row)] = round(df.loc[(df['Geography'] == 'Block') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population for Urban Blocks"), 'MeanPercDiff'].iat[0], 2)
ws['H'+str(row)] = round(df.loc[(df['Geography'] == 'Block') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population for Urban Blocks"), 'NumberBtw2Perc5Perc'].iat[0], 2)
ws['I'+str(row)] = round(df.loc[(df['Geography'] == 'Block') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population for Urban Blocks"), 'NumberGreater5Perc'].iat[0], 2)
ws['J'+str(row)] = round(df.loc[(df['Geography'] == 'Block') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population for Urban Blocks"), 'NumberGreater200'].iat[0], 2)
ws['B'+str(row+1)] = df.loc[(df['Geography'] == 'Block') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population for Rural Blocks"), 'NumCells'].iat[0]
ws['C'+str(row+1)] = round(df.loc[(df['Geography'] == 'Block') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population for Rural Blocks"), 'MeanAbsDiff'].iat[0], 2)
ws['D'+str(row+1)] = round(df.loc[(df['Geography'] == 'Block') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population for Rural Blocks"), 'RMSE'].iat[0], 2)
ws['E'+str(row+1)] = round(df.loc[(df['Geography'] == 'Block') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population for Rural Blocks"), 'MeanAbsPercDiff'].iat[0], 2)
ws['F'+str(row+1)] = round(df.loc[(df['Geography'] == 'Block') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population for Rural Blocks"), 'CV'].iat[0], 2)
ws['G'+str(row+1)] = round(df.loc[(df['Geography'] == 'Block') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population for Rural Blocks"), 'MeanPercDiff'].iat[0], 2)
ws['H'+str(row+1)] = round(df.loc[(df['Geography'] == 'Block') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population for Rural Blocks"), 'NumberBtw2Perc5Perc'].iat[0], 2)
ws['I'+str(row+1)] = round(df.loc[(df['Geography'] == 'Block') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population for Rural Blocks"), 'NumberGreater5Perc'].iat[0], 2)
ws['J'+str(row+1)] = round(df.loc[(df['Geography'] == 'Block') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population for Rural Blocks"), 'NumberGreater200'].iat[0], 2)

if prmetricsready:
    # Table 1e
    row = 45
    ws['B'+str(row)] = df.loc[(df['Geography'] == 'PR County/Municipio') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population"), 'NumCells'].iat[0]
    ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'PR County/Municipio') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population"), 'MeanAbsDiff'].iat[0], 2)
    ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'PR County/Municipio') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population"), 'RMSE'].iat[0], 2)
    ws['E'+str(row)] = round(df.loc[(df['Geography'] == 'PR County/Municipio') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population"), 'MeanAbsPercDiff'].iat[0], 2)
    ws['F'+str(row)] = round(df.loc[(df['Geography'] == 'PR County/Municipio') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population"), 'CV'].iat[0], 2)
    ws['G'+str(row)] = round(df.loc[(df['Geography'] == 'PR County/Municipio') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population"), 'MeanPercDiff'].iat[0], 2)
    ws['H'+str(row)] = df.loc[(df['Geography'] == 'PR County/Municipio') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population"), 'NumberBtw2Perc5Perc'].iat[0]
    ws['I'+str(row)] = df.loc[(df['Geography'] == 'PR County/Municipio') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population"), 'NumberGreater5Perc'].iat[0]
    ws['J'+str(row)] = df.loc[(df['Geography'] == 'PR County/Municipio') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population"), 'NumberGreater200'].iat[0]

    # Table 1f
    row = 51
    ws['B'+str(row)] = df.loc[(df['Geography'] == 'PR Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population"), 'NumCells'].iat[0]
    ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'PR Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population"), 'MeanAbsDiff'].iat[0], 2)
    ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'PR Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population"), 'RMSE'].iat[0], 2)
    ws['E'+str(row)] = df.loc[(df['Geography'] == 'PR Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population"), 'NumberBtw2Perc5Perc'].iat[0]
    ws['F'+str(row)] = df.loc[(df['Geography'] == 'PR Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population"), 'NumberGreater5Perc'].iat[0]
    ws['G'+str(row)] = df.loc[(df['Geography'] == 'PR Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population"), 'NumberGreater200'].iat[0]

# Table 1g 
row = 57
for s in sizecats_plusall:
    ws['B'+str(row)] = df.loc[(df['Geography'] == 'ESD') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'NumCells'].iat[0]
    ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'ESD') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'MeanAbsDiff'].iat[0], 2)
    ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'ESD') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'RMSE'].iat[0], 2)
    ws['E'+str(row)] = round(df.loc[(df['Geography'] == 'ESD') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'MeanAbsPercDiff'].iat[0], 2)
    ws['F'+str(row)] = round(df.loc[(df['Geography'] == 'ESD') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'CV'].iat[0], 2)
    ws['G'+str(row)] = round(df.loc[(df['Geography'] == 'ESD') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'MeanPercDiff'].iat[0], 2)
    ws['H'+str(row)] = df.loc[(df['Geography'] == 'ESD') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'NumberBtw2Perc5Perc'].iat[0]
    ws['I'+str(row)] = df.loc[(df['Geography'] == 'ESD') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'NumberGreater5Perc'].iat[0]
    ws['J'+str(row)] = df.loc[(df['Geography'] == 'ESD') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'NumberGreater200'].iat[0]
    row += 1
row = 65
for s in sizecats_plusall:
    ws['B'+str(row)] = df.loc[(df['Geography'] == 'SSD') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'NumCells'].iat[0]
    ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'SSD') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'MeanAbsDiff'].iat[0], 2)
    ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'SSD') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'RMSE'].iat[0], 2)
    ws['E'+str(row)] = round(df.loc[(df['Geography'] == 'SSD') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'MeanAbsPercDiff'].iat[0], 2)
    ws['F'+str(row)] = round(df.loc[(df['Geography'] == 'SSD') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'CV'].iat[0], 2)
    ws['G'+str(row)] = round(df.loc[(df['Geography'] == 'SSD') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'MeanPercDiff'].iat[0], 2)
    ws['H'+str(row)] = df.loc[(df['Geography'] == 'SSD') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'NumberBtw2Perc5Perc'].iat[0]
    ws['I'+str(row)] = df.loc[(df['Geography'] == 'SSD') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'NumberGreater5Perc'].iat[0]
    ws['J'+str(row)] = df.loc[(df['Geography'] == 'SSD') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'NumberGreater200'].iat[0]
    row += 1
row = 73
for s in sizecats_plusall:
    ws['B'+str(row)] = df.loc[(df['Geography'] == 'USD') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'NumCells'].iat[0]
    ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'USD') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'MeanAbsDiff'].iat[0], 2)
    ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'USD') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'RMSE'].iat[0], 2)
    ws['E'+str(row)] = round(df.loc[(df['Geography'] == 'USD') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'MeanAbsPercDiff'].iat[0], 2)
    ws['F'+str(row)] = round(df.loc[(df['Geography'] == 'USD') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'CV'].iat[0], 2)
    ws['G'+str(row)] = round(df.loc[(df['Geography'] == 'USD') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'MeanPercDiff'].iat[0], 2)
    ws['H'+str(row)] = df.loc[(df['Geography'] == 'USD') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'NumberBtw2Perc5Perc'].iat[0]
    ws['I'+str(row)] = df.loc[(df['Geography'] == 'USD') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'NumberGreater5Perc'].iat[0]
    ws['J'+str(row)] = df.loc[(df['Geography'] == 'USD') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'NumberGreater200'].iat[0]
    row += 1

# Table 1h
row = 85
for s in sizecats_plusall:
    ws['B'+str(row)] = df.loc[(df['Geography'] == 'MCD') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'NumCells'].iat[0]
    ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'MCD') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'MeanAbsDiff'].iat[0], 2)
    ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'MCD') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'RMSE'].iat[0], 2)
    ws['E'+str(row)] = round(df.loc[(df['Geography'] == 'MCD') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'MeanAbsPercDiff'].iat[0], 2)
    ws['F'+str(row)] = round(df.loc[(df['Geography'] == 'MCD') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'CV'].iat[0], 2)
    ws['G'+str(row)] = round(df.loc[(df['Geography'] == 'MCD') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'MeanPercDiff'].iat[0], 2)
    ws['H'+str(row)] = df.loc[(df['Geography'] == 'MCD') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'NumberBtw2Perc5Perc'].iat[0]
    ws['I'+str(row)] = df.loc[(df['Geography'] == 'MCD') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'NumberGreater5Perc'].iat[0]
    ws['J'+str(row)] = df.loc[(df['Geography'] == 'MCD') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'NumberGreater200'].iat[0]
    row += 1

# Table 1i
row = 97
for s in airanvsasizecats_plusall:
    ws['B'+str(row)] = df.loc[(df['Geography'] == 'Fed AIR') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'NumCells'].iat[0]
    ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'Fed AIR') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'MeanAbsDiff'].iat[0], 2)
    ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'Fed AIR') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'RMSE'].iat[0], 2)
    ws['E'+str(row)] = round(df.loc[(df['Geography'] == 'Fed AIR') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'MeanAbsPercDiff'].iat[0], 2)
    ws['F'+str(row)] = round(df.loc[(df['Geography'] == 'Fed AIR') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'CV'].iat[0], 2)
    ws['G'+str(row)] = round(df.loc[(df['Geography'] == 'Fed AIR') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'MeanPercDiff'].iat[0], 2)
    ws['H'+str(row)] = df.loc[(df['Geography'] == 'Fed AIR') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'NumberBtw2Perc5Perc'].iat[0]
    ws['I'+str(row)] = df.loc[(df['Geography'] == 'Fed AIR') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'NumberGreater5Perc'].iat[0]
    ws['J'+str(row)] = df.loc[(df['Geography'] == 'Fed AIR') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'NumberGreater200'].iat[0]
    row += 1

# Table 1j
row = 107
ws['B'+str(row)] = df.loc[(df['Geography'] == 'OTSA') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population"), 'NumCells'].iat[0]
ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'OTSA') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population"), 'MeanAbsDiff'].iat[0], 2)
ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'OTSA') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population"), 'RMSE'].iat[0], 2)
ws['E'+str(row)] = round(df.loc[(df['Geography'] == 'OTSA') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population"), 'MeanAbsPercDiff'].iat[0], 2)
ws['F'+str(row)] = round(df.loc[(df['Geography'] == 'OTSA') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population"), 'CV'].iat[0], 2)
ws['G'+str(row)] = round(df.loc[(df['Geography'] == 'OTSA') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population"), 'MeanPercDiff'].iat[0], 2)
ws['H'+str(row)] = df.loc[(df['Geography'] == 'OTSA') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population"), 'NumberBtw2Perc5Perc'].iat[0]
ws['I'+str(row)] = df.loc[(df['Geography'] == 'OTSA') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population"), 'NumberGreater5Perc'].iat[0]
ws['J'+str(row)] = df.loc[(df['Geography'] == 'OTSA') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population"), 'NumberGreater200'].iat[0]

# Table 1k
row = 113
for s in airanvsasizecats_plusall:
    ws['B'+str(row)] = df.loc[(df['Geography'] == 'ANVSA') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'NumCells'].iat[0]
    ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'ANVSA') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'MeanAbsDiff'].iat[0], 2)
    ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'ANVSA') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'RMSE'].iat[0], 2)
    ws['E'+str(row)] = round(df.loc[(df['Geography'] == 'ANVSA') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'MeanAbsPercDiff'].iat[0], 2)
    ws['F'+str(row)] = round(df.loc[(df['Geography'] == 'ANVSA') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'CV'].iat[0], 2)
    ws['G'+str(row)] = round(df.loc[(df['Geography'] == 'ANVSA') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'MeanPercDiff'].iat[0], 2)
    ws['H'+str(row)] = df.loc[(df['Geography'] == 'ANVSA') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'NumberBtw2Perc5Perc'].iat[0]
    ws['I'+str(row)] = df.loc[(df['Geography'] == 'ANVSA') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'NumberGreater5Perc'].iat[0]
    ws['J'+str(row)] = df.loc[(df['Geography'] == 'ANVSA') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population"), 'NumberGreater200'].iat[0]
    row += 1


# Table 2a
row = 123
for s in sizecats_plusall:
    ws['B'+str(row)] = df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population Aged 18+"), 'NumCells'].iat[0]
    ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population Aged 18+"), 'MeanAbsDiff'].iat[0], 2)
    ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population Aged 18+"), 'RMSE'].iat[0], 2)
    ws['E'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population Aged 18+"), 'MeanAbsPercDiff'].iat[0], 2)
    ws['F'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population Aged 18+"), 'CV'].iat[0], 2)
    ws['G'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population Aged 18+"), 'MeanPercDiff'].iat[0], 2)
    ws['H'+str(row)] = df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population Aged 18+"), 'NumberBtw2Perc5Perc'].iat[0]
    ws['I'+str(row)] = df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population Aged 18+"), 'NumberGreater5Perc'].iat[0]
    ws['J'+str(row)] = df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population Aged 18+"), 'NumberGreater200'].iat[0]
    row += 1

## Table 2b
row = 135
for s in incplacesizecats_plusall:
    ws['B'+str(row)] = df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population Aged 18+"), 'NumCells'].iat[0]
    ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population Aged 18+"), 'MeanAbsDiff'].iat[0], 2)
    ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population Aged 18+"), 'RMSE'].iat[0], 2)
    ws['E'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population Aged 18+"), 'MeanAbsPercDiff'].iat[0], 2)
    ws['F'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population Aged 18+"), 'CV'].iat[0], 2)
    ws['G'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population Aged 18+"), 'MeanPercDiff'].iat[0], 2)
    ws['H'+str(row)] = df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population Aged 18+"), 'NumberBtw2Perc5Perc'].iat[0]
    ws['I'+str(row)] = df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population Aged 18+"), 'NumberGreater5Perc'].iat[0]
    ws['J'+str(row)] = df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population Aged 18+"), 'NumberGreater200'].iat[0]
    row += 1

# Table 2c
row = 148
ws['B'+str(row)] = df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population Aged 18+"), 'NumCells'].iat[0]
ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population Aged 18+"), 'MeanAbsDiff'].iat[0], 2)
ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population Aged 18+"), 'RMSE'].iat[0], 2)
ws['E'+str(row)] = df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population Aged 18+"), 'NumberBtw2Perc5Perc'].iat[0]
ws['F'+str(row)] = df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population Aged 18+"), 'NumberGreater5Perc'].iat[0]
ws['G'+str(row)] = df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population Aged 18+"), 'NumberGreater200'].iat[0]

if prmetricsready:
    # Table 2d
    row = 154
    ws['B'+str(row)] = df.loc[(df['Geography'] == 'PR County/Municipio') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population Aged 18+"), 'NumCells'].iat[0]
    ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'PR County/Municipio') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population Aged 18+"), 'MeanAbsDiff'].iat[0], 2)
    ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'PR County/Municipio') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population Aged 18+"), 'RMSE'].iat[0], 2)
    ws['E'+str(row)] = round(df.loc[(df['Geography'] == 'PR County/Municipio') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population Aged 18+"), 'MeanAbsPercDiff'].iat[0], 2)
    ws['F'+str(row)] = round(df.loc[(df['Geography'] == 'PR County/Municipio') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population Aged 18+"), 'CV'].iat[0], 2)
    ws['G'+str(row)] = round(df.loc[(df['Geography'] == 'PR County/Municipio') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population Aged 18+"), 'MeanPercDiff'].iat[0], 2)
    ws['H'+str(row)] = df.loc[(df['Geography'] == 'PR County/Municipio') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population Aged 18+"), 'NumberBtw2Perc5Perc'].iat[0]
    ws['I'+str(row)] = df.loc[(df['Geography'] == 'PR County/Municipio') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population Aged 18+"), 'NumberGreater5Perc'].iat[0]
    ws['J'+str(row)] = df.loc[(df['Geography'] == 'PR County/Municipio') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population Aged 18+"), 'NumberGreater200'].iat[0]

    # Table 2e
    row = 160
    ws['B'+str(row)] = df.loc[(df['Geography'] == 'PR Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population Aged 18+"), 'NumCells'].iat[0]
    ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'PR Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population Aged 18+"), 'MeanAbsDiff'].iat[0], 2)
    ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'PR Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population Aged 18+"), 'RMSE'].iat[0], 2)
    ws['E'+str(row)] = df.loc[(df['Geography'] == 'PR Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population Aged 18+"), 'NumberBtw2Perc5Perc'].iat[0]
    ws['F'+str(row)] = df.loc[(df['Geography'] == 'PR Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population Aged 18+"), 'NumberGreater5Perc'].iat[0]
    ws['G'+str(row)] = df.loc[(df['Geography'] == 'PR Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population Aged 18+"), 'NumberGreater200'].iat[0]

# Table 2f
row = 166
for s in sizecats_plusall:
    ws['B'+str(row)] = df.loc[(df['Geography'] == 'MCD') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population Aged 18+"), 'NumCells'].iat[0]
    ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'MCD') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population Aged 18+"), 'MeanAbsDiff'].iat[0], 2)
    ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'MCD') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population Aged 18+"), 'RMSE'].iat[0], 2)
    ws['E'+str(row)] = round(df.loc[(df['Geography'] == 'MCD') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population Aged 18+"), 'MeanAbsPercDiff'].iat[0], 2)
    ws['F'+str(row)] = round(df.loc[(df['Geography'] == 'MCD') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population Aged 18+"), 'CV'].iat[0], 2)
    ws['G'+str(row)] = round(df.loc[(df['Geography'] == 'MCD') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population Aged 18+"), 'MeanPercDiff'].iat[0], 2)
    ws['H'+str(row)] = df.loc[(df['Geography'] == 'MCD') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population Aged 18+"), 'NumberBtw2Perc5Perc'].iat[0]
    ws['I'+str(row)] = df.loc[(df['Geography'] == 'MCD') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population Aged 18+"), 'NumberGreater5Perc'].iat[0]
    ws['J'+str(row)] = df.loc[(df['Geography'] == 'MCD') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population Aged 18+"), 'NumberGreater200'].iat[0]
    row += 1

# Table 2g
row = 178
for s in airanvsasizecats_plusall:
    ws['B'+str(row)] = df.loc[(df['Geography'] == 'Fed AIR') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population Aged 18+"), 'NumCells'].iat[0]
    ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'Fed AIR') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population Aged 18+"), 'MeanAbsDiff'].iat[0], 2)
    ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'Fed AIR') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population Aged 18+"), 'RMSE'].iat[0], 2)
    ws['E'+str(row)] = round(df.loc[(df['Geography'] == 'Fed AIR') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population Aged 18+"), 'MeanAbsPercDiff'].iat[0], 2)
    ws['F'+str(row)] = round(df.loc[(df['Geography'] == 'Fed AIR') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population Aged 18+"), 'CV'].iat[0], 2)
    ws['G'+str(row)] = round(df.loc[(df['Geography'] == 'Fed AIR') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population Aged 18+"), 'MeanPercDiff'].iat[0], 2)
    ws['H'+str(row)] = df.loc[(df['Geography'] == 'Fed AIR') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population Aged 18+"), 'NumberBtw2Perc5Perc'].iat[0]
    ws['I'+str(row)] = df.loc[(df['Geography'] == 'Fed AIR') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population Aged 18+"), 'NumberGreater5Perc'].iat[0]
    ws['J'+str(row)] = df.loc[(df['Geography'] == 'Fed AIR') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population Aged 18+"), 'NumberGreater200'].iat[0]
    row += 1

# Table 2h
row = 188
ws['B'+str(row)] = df.loc[(df['Geography'] == 'OTSA') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population Aged 18+"), 'NumCells'].iat[0]
ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'OTSA') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population Aged 18+"), 'MeanAbsDiff'].iat[0], 2)
ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'OTSA') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population Aged 18+"), 'RMSE'].iat[0], 2)
ws['E'+str(row)] = round(df.loc[(df['Geography'] == 'OTSA') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population Aged 18+"), 'MeanAbsPercDiff'].iat[0], 2)
ws['F'+str(row)] = round(df.loc[(df['Geography'] == 'OTSA') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population Aged 18+"), 'CV'].iat[0], 2)
ws['G'+str(row)] = round(df.loc[(df['Geography'] == 'OTSA') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population Aged 18+"), 'MeanPercDiff'].iat[0], 2)
ws['H'+str(row)] = df.loc[(df['Geography'] == 'OTSA') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population Aged 18+"), 'NumberBtw2Perc5Perc'].iat[0]
ws['I'+str(row)] = df.loc[(df['Geography'] == 'OTSA') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population Aged 18+"), 'NumberGreater5Perc'].iat[0]
ws['J'+str(row)] = df.loc[(df['Geography'] == 'OTSA') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population Aged 18+"), 'NumberGreater200'].iat[0]

# Table 2i
row = 194
for s in airanvsasizecats_plusall:
    ws['B'+str(row)] = df.loc[(df['Geography'] == 'ANVSA') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population Aged 18+"), 'NumCells'].iat[0]
    ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'ANVSA') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population Aged 18+"), 'MeanAbsDiff'].iat[0], 2)
    ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'ANVSA') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population Aged 18+"), 'RMSE'].iat[0], 2)
    ws['E'+str(row)] = round(df.loc[(df['Geography'] == 'ANVSA') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population Aged 18+"), 'MeanAbsPercDiff'].iat[0], 2)
    ws['F'+str(row)] = round(df.loc[(df['Geography'] == 'ANVSA') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population Aged 18+"), 'CV'].iat[0], 2)
    ws['G'+str(row)] = round(df.loc[(df['Geography'] == 'ANVSA') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population Aged 18+"), 'MeanPercDiff'].iat[0], 2)
    ws['H'+str(row)] = df.loc[(df['Geography'] == 'ANVSA') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population Aged 18+"), 'NumberBtw2Perc5Perc'].iat[0]
    ws['I'+str(row)] = df.loc[(df['Geography'] == 'ANVSA') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population Aged 18+"), 'NumberGreater5Perc'].iat[0]
    ws['J'+str(row)] = df.loc[(df['Geography'] == 'ANVSA') & (df['Size_Category'] == s) & (df['Characteristic'] == "Total Population Aged 18+"), 'NumberGreater200'].iat[0]
    row += 1

if unitmetricsready:
    # Table 3a
    row = 204
    ws['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumCells'].iat[0]
    ws['C'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'MeanAbsDiff'].iat[0], 2)
    ws['D'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'RMSE'].iat[0], 2)
    ws['E'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'MeanDiff'].iat[0], 2)
    ws['F'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumberHundredOnMDFNotHDF'].iat[0])
    ws['G'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumberZeroOnMDFNotHDF'].iat[0])
    ws['H'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumberBtw2Pts5Pts'].iat[0])
    ws['I'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumberGreater5Pts'].iat[0])
    # Table 3b 
    row = 210
    ws['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumCells'].iat[0]
    ws['C'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'MeanAbsDiff'].iat[0], 2)
    ws['D'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'RMSE'].iat[0], 2)
    ws['E'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'MeanDiff'].iat[0], 2)
    ws['F'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumberHundredOnMDFNotHDF'].iat[0])
    ws['G'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumberZeroOnMDFNotHDF'].iat[0])
    ws['H'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumberBtw2Pts5Pts'].iat[0])
    ws['I'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumberGreater5Pts'].iat[0])
    # Table 3c
    row = 216
    ws['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumCells'].iat[0]
    ws['C'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'MeanAbsDiff'].iat[0], 2)
    ws['D'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'RMSE'].iat[0], 2)
    ws['E'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumberHundredOnMDFNotHDF'].iat[0])
    ws['F'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumberZeroOnMDFNotHDF'].iat[0])
    ws['G'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumberBtw2Pts5Pts'].iat[0])
    ws['H'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumberGreater5Pts'].iat[0])
    if prunitmetricsready:
        # Table 3d
        row = 222
        ws['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'PR County/Municipio') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumCells'].iat[0]
        ws['C'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'PR County/Municipio') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'MeanAbsDiff'].iat[0], 2)
        ws['D'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'PR County/Municipio') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'RMSE'].iat[0], 2)
        ws['E'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'PR County/Municipio') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'MeanDiff'].iat[0], 2)
        ws['F'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'PR County/Municipio') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumberHundredOnMDFNotHDF'].iat[0])
        ws['G'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'PR County/Municipio') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumberZeroOnMDFNotHDF'].iat[0])
        ws['H'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'PR County/Municipio') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumberBtw2Pts5Pts'].iat[0])
        ws['I'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'PR County/Municipio') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumberGreater5Pts'].iat[0])
        # Table 3e
        row = 228
        ws['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'PR Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumCells'].iat[0]
        ws['C'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'PR Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'MeanAbsDiff'].iat[0], 2)
        ws['D'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'PR Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'RMSE'].iat[0], 2)
        ws['E'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'PR Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumberHundredOnMDFNotHDF'].iat[0])
        ws['F'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'PR Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumberZeroOnMDFNotHDF'].iat[0])
        ws['G'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'PR Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumberBtw2Pts5Pts'].iat[0])
        ws['H'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'PR Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumberGreater5Pts'].iat[0])
    # Table 3f
    row = 234
    ws['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'ESD') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumCells'].iat[0]
    ws['C'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'ESD') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'MeanAbsDiff'].iat[0], 2)
    ws['D'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'ESD') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'RMSE'].iat[0], 2)
    ws['E'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'ESD') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'MeanDiff'].iat[0], 2)
    ws['F'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'ESD') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumberHundredOnMDFNotHDF'].iat[0])
    ws['G'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'ESD') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumberZeroOnMDFNotHDF'].iat[0])
    ws['H'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'ESD') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumberBtw2Pts5Pts'].iat[0])
    ws['I'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'ESD') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumberGreater5Pts'].iat[0])
    ws['B'+str(row+1)] = dfu.loc[(dfu['Geography'] == 'SSD') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumCells'].iat[0]
    ws['C'+str(row+1)] = round(dfu.loc[(dfu['Geography'] == 'SSD') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'MeanAbsDiff'].iat[0], 2)
    ws['D'+str(row+1)] = round(dfu.loc[(dfu['Geography'] == 'SSD') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'RMSE'].iat[0], 2)
    ws['E'+str(row+1)] = round(dfu.loc[(dfu['Geography'] == 'SSD') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'MeanDiff'].iat[0], 2)
    ws['F'+str(row+1)] = round(dfu.loc[(dfu['Geography'] == 'SSD') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumberHundredOnMDFNotHDF'].iat[0])
    ws['G'+str(row+1)] = round(dfu.loc[(dfu['Geography'] == 'SSD') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumberZeroOnMDFNotHDF'].iat[0])
    ws['H'+str(row+1)] = round(dfu.loc[(dfu['Geography'] == 'SSD') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumberBtw2Pts5Pts'].iat[0])
    ws['I'+str(row+1)] = round(dfu.loc[(dfu['Geography'] == 'SSD') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumberGreater5Pts'].iat[0])  
    ws['B'+str(row+2)] = dfu.loc[(dfu['Geography'] == 'USD') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumCells'].iat[0]
    ws['C'+str(row+2)] = round(dfu.loc[(dfu['Geography'] == 'USD') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'MeanAbsDiff'].iat[0], 2)
    ws['D'+str(row+2)] = round(dfu.loc[(dfu['Geography'] == 'USD') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'RMSE'].iat[0], 2)
    ws['E'+str(row+2)] = round(dfu.loc[(dfu['Geography'] == 'USD') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'MeanDiff'].iat[0], 2)
    ws['F'+str(row+2)] = round(dfu.loc[(dfu['Geography'] == 'USD') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumberHundredOnMDFNotHDF'].iat[0])
    ws['G'+str(row+2)] = round(dfu.loc[(dfu['Geography'] == 'USD') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumberZeroOnMDFNotHDF'].iat[0])
    ws['H'+str(row+2)] = round(dfu.loc[(dfu['Geography'] == 'USD') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumberBtw2Pts5Pts'].iat[0])
    ws['I'+str(row+2)] = round(dfu.loc[(dfu['Geography'] == 'USD') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumberGreater5Pts'].iat[0])
    # Table 3g
    row = 242
    ws['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'MCD') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumCells'].iat[0]
    ws['C'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'MCD') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'MeanAbsDiff'].iat[0], 2)
    ws['D'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'MCD') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'RMSE'].iat[0], 2)
    ws['E'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'MCD') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'MeanDiff'].iat[0], 2)
    ws['F'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'MCD') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumberHundredOnMDFNotHDF'].iat[0])
    ws['G'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'MCD') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumberZeroOnMDFNotHDF'].iat[0])
    ws['H'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'MCD') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumberBtw2Pts5Pts'].iat[0])
    ws['I'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'MCD') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumberGreater5Pts'].iat[0])
    # Table 3h
    row = 248
    ws['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'Fed AIR') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumCells'].iat[0]
    ws['C'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Fed AIR') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'MeanAbsDiff'].iat[0], 2)
    ws['D'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Fed AIR') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'RMSE'].iat[0], 2)
    ws['E'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Fed AIR') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'MeanDiff'].iat[0], 2)
    ws['F'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Fed AIR') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumberHundredOnMDFNotHDF'].iat[0])
    ws['G'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Fed AIR') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumberZeroOnMDFNotHDF'].iat[0])
    ws['H'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Fed AIR') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumberBtw2Pts5Pts'].iat[0])
    ws['I'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Fed AIR') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumberGreater5Pts'].iat[0])
    # Table 3i
    row = 254
    ws['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'OTSA') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumCells'].iat[0]
    ws['C'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'OTSA') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'MeanAbsDiff'].iat[0], 2)
    ws['D'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'OTSA') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'RMSE'].iat[0], 2)
    ws['E'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'OTSA') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'MeanDiff'].iat[0], 2)
    ws['F'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'OTSA') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumberHundredOnMDFNotHDF'].iat[0])
    ws['G'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'OTSA') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumberZeroOnMDFNotHDF'].iat[0])
    ws['H'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'OTSA') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumberBtw2Pts5Pts'].iat[0])
    ws['I'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'OTSA') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumberGreater5Pts'].iat[0])
    # Table 3j
    row = 260
    ws['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'ANVSA') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumCells'].iat[0]
    ws['C'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'ANVSA') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'MeanAbsDiff'].iat[0], 2)
    ws['D'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'ANVSA') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'RMSE'].iat[0], 2)
    ws['E'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'ANVSA') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'MeanDiff'].iat[0], 2)
    ws['F'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'ANVSA') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumberHundredOnMDFNotHDF'].iat[0])
    ws['G'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'ANVSA') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumberZeroOnMDFNotHDF'].iat[0])
    ws['H'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'ANVSA') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumberBtw2Pts5Pts'].iat[0])
    ws['I'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'ANVSA') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumberGreater5Pts'].iat[0])
    # Table 4a
    row = 266
    ws['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "PPH"), 'NumCells'].iat[0]
    ws['C'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "PPH"), 'MedianAbsDiff'].iat[0], 2)
    ws['D'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "PPH"), 'RMSE'].iat[0], 2)
    ws['E'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "PPH"), 'MedianDiff'].iat[0], 2)
    ws['F'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "PPH"), 'NumberBtw2Perc5Perc'].iat[0])
    ws['G'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "PPH"), 'NumberGreater5Perc'].iat[0])
    # Table 4b
    row = 273
    ws['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "PPH"), 'NumCells'].iat[0]
    ws['C'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "PPH"), 'MedianAbsDiff'].iat[0], 2)
    ws['D'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "PPH"), 'RMSE'].iat[0], 2)
    ws['E'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "PPH"), 'MedianDiff'].iat[0], 2)
    ws['F'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "PPH"), 'NumberBtw2Perc5Perc'].iat[0])
    ws['G'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "PPH"), 'NumberGreater5Perc'].iat[0])
    # Table 4c
    row = 280
    ws['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "PPH"), 'NumCells'].iat[0]
    ws['C'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "PPH"), 'MedianAbsDiff'].iat[0], 2)
    ws['D'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "PPH"), 'RMSE'].iat[0], 2)
    ws['E'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "PPH"), 'MedianDiff'].iat[0], 2)
    ws['F'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "PPH"), 'NumberBtw2Perc5Perc'].iat[0])
    ws['G'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "PPH"), 'NumberGreater5Perc'].iat[0])
    if prunitmetricsready:
        # Table 4d
        row = 287
        ws['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'PR County/Municipio') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "PPH"), 'NumCells'].iat[0]
        ws['C'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'PR County/Municipio') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "PPH"), 'MedianAbsDiff'].iat[0], 2)
        ws['D'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'PR County/Municipio') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "PPH"), 'RMSE'].iat[0], 2)
        ws['E'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'PR County/Municipio') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "PPH"), 'MedianDiff'].iat[0], 2)
        ws['F'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'PR County/Municipio') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "PPH"), 'NumberBtw2Perc5Perc'].iat[0])
        ws['G'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'PR County/Municipio') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "PPH"), 'NumberGreater5Perc'].iat[0])
        # Table 4e
        row = 293
        ws['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'PR Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "PPH"), 'NumCells'].iat[0]
        ws['C'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'PR Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "PPH"), 'MedianAbsDiff'].iat[0], 2)
        ws['D'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'PR Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "PPH"), 'RMSE'].iat[0], 2)
        ws['E'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'PR Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "PPH"), 'MedianDiff'].iat[0], 2)
        ws['F'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'PR Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "PPH"), 'NumberBtw2Perc5Perc'].iat[0])
        ws['G'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'PR Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "PPH"), 'NumberGreater5Perc'].iat[0])

# Table 5a
row = 300
for c in ['Hispanic', 'Not Hispanic']:
    ws['B'+str(row)] = df.loc[(df['Geography'] == 'State') & (df['Size_Category'] == 'All') & (df['Characteristic'] == c), 'NumCells'].iat[0]
    ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'State') & (df['Size_Category'] == 'All') & (df['Characteristic'] == c), 'MeanAbsDiff'].iat[0], 2)
    ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'State') & (df['Size_Category'] == 'All') & (df['Characteristic'] == c), 'RMSE'].iat[0], 2)
    ws['E'+str(row)] = round(df.loc[(df['Geography'] == 'State') & (df['Size_Category'] == 'All') & (df['Characteristic'] == c), 'MeanAbsPercDiff'].iat[0], 2)
    ws['F'+str(row)] = round(df.loc[(df['Geography'] == 'State') & (df['Size_Category'] == 'All') & (df['Characteristic'] == c), 'CV'].iat[0], 2)
    ws['G'+str(row)] = round(df.loc[(df['Geography'] == 'State') & (df['Size_Category'] == 'All') & (df['Characteristic'] == c), 'MeanPercDiff'].iat[0], 2)
    ws['H'+str(row)] = round(df.loc[(df['Geography'] == 'State') & (df['Size_Category'] == 'All') & (df['Characteristic'] == c), 'NumberGreater10Perc'].iat[0], 2)
    row += 1

# Table 5b
row = 307
for s in smallsizecats_plusall:
    row += 1
    for c in ['Hispanic', 'Not Hispanic']:
        ws['B'+str(row)] = df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == c), 'NumCells'].iat[0]
        ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == c), 'MeanAbsDiff'].iat[0], 2)
        ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == c), 'RMSE'].iat[0], 2)
        ws['E'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == c), 'MeanAbsPercDiff'].iat[0], 2)
        ws['F'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == c), 'CV'].iat[0], 2)
        ws['G'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == c), 'MeanPercDiff'].iat[0], 2)
        ws['H'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == c), 'NumberGreater10Perc'].iat[0], 2)
        row += 1

# Table 5c
row = 324
for s in smallsizecats_plusall:
    row += 1
    for c in ['Hispanic', 'Not Hispanic']:
        ws['B'+str(row)] = df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == c), 'NumCells'].iat[0]
        ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == c), 'MeanAbsDiff'].iat[0], 2)
        ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == c), 'RMSE'].iat[0], 2)
        ws['E'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == c), 'MeanAbsPercDiff'].iat[0], 2)
        ws['F'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == c), 'CV'].iat[0], 2)
        ws['G'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == c), 'MeanPercDiff'].iat[0], 2)
        ws['H'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == c), 'NumberGreater10Perc'].iat[0], 2)
        row += 1

# Table 5d
row = 341
for s in smallsizecats_plusall:
    row += 1
    for c in ['Hispanic', 'Not Hispanic']:
        ws['B'+str(row)] = df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == s) & (df['Characteristic'] == c), 'NumCells'].iat[0]
        ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == s) & (df['Characteristic'] == c), 'MeanAbsDiff'].iat[0], 2)
        ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == s) & (df['Characteristic'] == c), 'RMSE'].iat[0], 2)
        ws['E'+str(row)] = round(df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == s) & (df['Characteristic'] == c), 'NumberGreater10Perc'].iat[0], 2)
        row += 1

# Table 6a
row = 358
for r in racealonecats:
    ws['B'+str(row)] = df.loc[(df['Geography'] == 'State') & (df['Size_Category'] == "All") & (df['Characteristic'] == r), 'NumCells'].iat[0]
    ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'State') & (df['Size_Category'] == "All") & (df['Characteristic'] == r), 'MeanAbsDiff'].iat[0], 2)
    ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'State') & (df['Size_Category'] == "All") & (df['Characteristic'] == r), 'RMSE'].iat[0], 2)
    ws['E'+str(row)] = round(df.loc[(df['Geography'] == 'State') & (df['Size_Category'] == "All") & (df['Characteristic'] == r), 'MeanAbsPercDiff'].iat[0], 2)
    ws['F'+str(row)] = round(df.loc[(df['Geography'] == 'State') & (df['Size_Category'] == "All") & (df['Characteristic'] == r), 'CV'].iat[0], 2)
    ws['G'+str(row)] = round(df.loc[(df['Geography'] == 'State') & (df['Size_Category'] == "All") & (df['Characteristic'] == r), 'MeanPercDiff'].iat[0], 2)
    ws['H'+str(row)] = df.loc[(df['Geography'] == 'State') & (df['Size_Category'] == "All") & (df['Characteristic'] == r), 'NumberGreater10Perc'].iat[0]
    row += 1

# Table 6b
row = 370
for s in smallsizecats_plusall:
    row += 1
    for r in racealonecats:
        ws['B'+str(row)] = df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'NumCells'].iat[0]
        ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'MeanAbsDiff'].iat[0], 2)
        ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'RMSE'].iat[0], 2)
        ws['E'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'MeanAbsPercDiff'].iat[0], 2)
        ws['F'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'CV'].iat[0], 2)
        ws['G'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'MeanPercDiff'].iat[0], 2)
        ws['H'+str(row)] = df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'NumberGreater10Perc'].iat[0]
        row += 1

# Table 6c
row = 407
for s in smallsizecats_plusall:
    row += 1
    for r in racealonecats:
        ws['B'+str(row)] = df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'NumCells'].iat[0]
        ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'MeanAbsDiff'].iat[0], 2)
        ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'RMSE'].iat[0], 2)
        ws['E'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'MeanAbsPercDiff'].iat[0], 2)
        ws['F'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'CV'].iat[0], 2)
        ws['G'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'MeanPercDiff'].iat[0], 2)
        ws['H'+str(row)] = df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'NumberGreater10Perc'].iat[0]
        row += 1

# Table 6d
row = 444
for s in smallsizecats_plusall:
    row += 1
    for r in racealonecats:
        ws['B'+str(row)] = df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'NumCells'].iat[0]
        ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'MeanAbsDiff'].iat[0], 2)
        ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'RMSE'].iat[0], 2)
        ws['E'+str(row)] = df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'NumberGreater10Perc'].iat[0]
        row += 1

if prmetricsready:
    # Table 6e
    row = 481
    for s in smallsizecats_plusall:
        row += 1
        for r in racealonecats:
            ws['B'+str(row)] = df.loc[(df['Geography'] == 'PR County/Municipio') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'NumCells'].iat[0]
            ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'PR County/Municipio') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'MeanAbsDiff'].iat[0], 2)
            ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'PR County/Municipio') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'RMSE'].iat[0], 2)
            ws['E'+str(row)] = round(df.loc[(df['Geography'] == 'PR County/Municipio') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'MeanAbsPercDiff'].iat[0], 2)
            ws['F'+str(row)] = round(df.loc[(df['Geography'] == 'PR County/Municipio') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'CV'].iat[0], 2)
            ws['G'+str(row)] = round(df.loc[(df['Geography'] == 'PR County/Municipio') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'MeanPercDiff'].iat[0], 2)
            ws['H'+str(row)] = df.loc[(df['Geography'] == 'PR County/Municipio') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'NumberGreater10Perc'].iat[0]
            row += 1

    # Table 6f
    row = 518 
    for s in smallsizecats_plusall:
        row += 1
        for r in racealonecats:
            ws['B'+str(row)] = df.loc[(df['Geography'] == 'PR Tract') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'NumCells'].iat[0]
            ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'PR Tract') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'MeanAbsDiff'].iat[0], 2)
            ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'PR Tract') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'RMSE'].iat[0], 2)
            ws['E'+str(row)] = df.loc[(df['Geography'] == 'PR Tract') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'NumberGreater10Perc'].iat[0]
            row += 1

# Table 6g
row = 555
for s in smallsizecats_plusall:
    row += 1
    for r in racealonecats:
        ws['B'+str(row)] = df.loc[(df['Geography'] == 'Fed AIR') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'NumCells'].iat[0]
        ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'Fed AIR') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'MeanAbsDiff'].iat[0], 2)
        ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'Fed AIR') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'RMSE'].iat[0], 2)
        ws['E'+str(row)] = round(df.loc[(df['Geography'] == 'Fed AIR') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'MeanAbsPercDiff'].iat[0], 2)
        ws['F'+str(row)] = round(df.loc[(df['Geography'] == 'Fed AIR') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'CV'].iat[0], 2)
        ws['G'+str(row)] = round(df.loc[(df['Geography'] == 'Fed AIR') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'MeanPercDiff'].iat[0], 2)
        ws['H'+str(row)] = df.loc[(df['Geography'] == 'Fed AIR') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'NumberGreater10Perc'].iat[0]
        row += 1

# Table 6h
row = 592
for s in smallsizecats_plusall:
    row += 1
    for r in racealonecats:
        ws['B'+str(row)] = df.loc[(df['Geography'] == 'OTSA') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'NumCells'].iat[0]
        ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'OTSA') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'MeanAbsDiff'].iat[0], 2)
        ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'OTSA') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'RMSE'].iat[0], 2)
        ws['E'+str(row)] = round(df.loc[(df['Geography'] == 'OTSA') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'MeanAbsPercDiff'].iat[0], 2)
        ws['F'+str(row)] = round(df.loc[(df['Geography'] == 'OTSA') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'CV'].iat[0], 2)
        ws['G'+str(row)] = round(df.loc[(df['Geography'] == 'OTSA') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'MeanPercDiff'].iat[0], 2)
        ws['H'+str(row)] = df.loc[(df['Geography'] == 'OTSA') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'NumberGreater10Perc'].iat[0]
        row += 1

# Table 6i
row = 629
for s in smallsizecats_plusall:
    row += 1
    for r in racealonecats:
        ws['B'+str(row)] = df.loc[(df['Geography'] == 'ANVSA') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'NumCells'].iat[0]
        ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'ANVSA') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'MeanAbsDiff'].iat[0], 2)
        ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'ANVSA') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'RMSE'].iat[0], 2)
        ws['E'+str(row)] = round(df.loc[(df['Geography'] == 'ANVSA') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'MeanAbsPercDiff'].iat[0], 2)
        ws['F'+str(row)] = round(df.loc[(df['Geography'] == 'ANVSA') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'CV'].iat[0], 2)
        ws['G'+str(row)] = round(df.loc[(df['Geography'] == 'ANVSA') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'MeanPercDiff'].iat[0], 2)
        ws['H'+str(row)] = df.loc[(df['Geography'] == 'ANVSA') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'NumberGreater10Perc'].iat[0]
        row += 1

# Table 7a
row = 666
for r in raceincombcats:
    ws['B'+str(row)] = df.loc[(df['Geography'] == 'State') & (df['Size_Category'] == "All") & (df['Characteristic'] == r), 'NumCells'].iat[0]
    ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'State') & (df['Size_Category'] == "All") & (df['Characteristic'] == r), 'MeanAbsDiff'].iat[0], 2)
    ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'State') & (df['Size_Category'] == "All") & (df['Characteristic'] == r), 'RMSE'].iat[0], 2)
    ws['E'+str(row)] = round(df.loc[(df['Geography'] == 'State') & (df['Size_Category'] == "All") & (df['Characteristic'] == r), 'MeanAbsPercDiff'].iat[0], 2)
    ws['F'+str(row)] = round(df.loc[(df['Geography'] == 'State') & (df['Size_Category'] == "All") & (df['Characteristic'] == r), 'CV'].iat[0], 2)
    ws['G'+str(row)] = round(df.loc[(df['Geography'] == 'State') & (df['Size_Category'] == "All") & (df['Characteristic'] == r), 'MeanPercDiff'].iat[0], 2)
    ws['H'+str(row)] = df.loc[(df['Geography'] == 'State') & (df['Size_Category'] == "All") & (df['Characteristic'] == r), 'NumberGreater10Perc'].iat[0]
    row += 1

# Table 7b
row = 677
for s in smallsizecats_plusall:
    row += 1
    for r in raceincombcats:
        ws['B'+str(row)] = df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'NumCells'].iat[0]
        ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'MeanAbsDiff'].iat[0], 2)
        ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'RMSE'].iat[0], 2)
        ws['E'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'MeanAbsPercDiff'].iat[0], 2)
        ws['F'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'CV'].iat[0], 2)
        ws['G'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'MeanPercDiff'].iat[0], 2)
        ws['H'+str(row)] = df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'NumberGreater10Perc'].iat[0]
        row += 1

# Table 7c
row = 710
for s in smallsizecats_plusall:
    row += 1
    for r in raceincombcats:
        ws['B'+str(row)] = df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'NumCells'].iat[0]
        ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'MeanAbsDiff'].iat[0], 2)
        ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'RMSE'].iat[0], 2)
        ws['E'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'MeanAbsPercDiff'].iat[0], 2)
        ws['F'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'CV'].iat[0], 2)
        ws['G'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'MeanPercDiff'].iat[0], 2)
        ws['H'+str(row)] = df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'NumberGreater10Perc'].iat[0]
        row += 1

# Table 7d
row = 743
for s in smallsizecats_plusall:
    row += 1
    for r in raceincombcats:
        ws['B'+str(row)] = df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'NumCells'].iat[0]
        ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'MeanAbsDiff'].iat[0], 2)
        ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'RMSE'].iat[0], 2)
        ws['E'+str(row)] = df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'NumberGreater10Perc'].iat[0]
        row += 1

if prmetricsready:
    # Table 7e
    row = 776
    for s in smallsizecats_plusall:
        row += 1
        for r in raceincombcats:
            ws['B'+str(row)] = df.loc[(df['Geography'] == 'PR County/Municipio') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'NumCells'].iat[0]
            ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'PR County/Municipio') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'MeanAbsDiff'].iat[0], 2)
            ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'PR County/Municipio') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'RMSE'].iat[0], 2)
            ws['E'+str(row)] = round(df.loc[(df['Geography'] == 'PR County/Municipio') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'MeanAbsPercDiff'].iat[0], 2)
            ws['F'+str(row)] = round(df.loc[(df['Geography'] == 'PR County/Municipio') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'CV'].iat[0], 2)
            ws['G'+str(row)] = round(df.loc[(df['Geography'] == 'PR County/Municipio') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'MeanPercDiff'].iat[0], 2)
            ws['H'+str(row)] = df.loc[(df['Geography'] == 'PR County/Municipio') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'NumberGreater10Perc'].iat[0]
            row += 1

    # Table 7f
    row = 809
    for s in smallsizecats_plusall:
        row += 1
        for r in raceincombcats:
            ws['B'+str(row)] = df.loc[(df['Geography'] == 'PR Tract') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'NumCells'].iat[0]
            ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'PR Tract') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'MeanAbsDiff'].iat[0], 2)
            ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'PR Tract') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'RMSE'].iat[0], 2)
            ws['E'+str(row)] = df.loc[(df['Geography'] == 'PR Tract') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'NumberGreater10Perc'].iat[0]
            row += 1

# Table 7g
row = 842
for s in smallsizecats_plusall:
    row += 1
    for r in raceincombcats:
        ws['B'+str(row)] = df.loc[(df['Geography'] == 'Fed AIR') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'NumCells'].iat[0]
        ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'Fed AIR') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'MeanAbsDiff'].iat[0], 2)
        ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'Fed AIR') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'RMSE'].iat[0], 2)
        ws['E'+str(row)] = round(df.loc[(df['Geography'] == 'Fed AIR') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'MeanAbsPercDiff'].iat[0], 2)
        ws['F'+str(row)] = round(df.loc[(df['Geography'] == 'Fed AIR') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'CV'].iat[0], 2)
        ws['G'+str(row)] = round(df.loc[(df['Geography'] == 'Fed AIR') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'MeanPercDiff'].iat[0], 2)
        ws['H'+str(row)] = df.loc[(df['Geography'] == 'Fed AIR') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'NumberGreater10Perc'].iat[0]
        row += 1

# Table 7h
row = 875
for s in smallsizecats_plusall:
    row += 1
    for r in raceincombcats:
        ws['B'+str(row)] = df.loc[(df['Geography'] == 'OTSA') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'NumCells'].iat[0]
        ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'OTSA') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'MeanAbsDiff'].iat[0], 2)
        ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'OTSA') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'RMSE'].iat[0], 2)
        ws['E'+str(row)] = round(df.loc[(df['Geography'] == 'OTSA') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'MeanAbsPercDiff'].iat[0], 2)
        ws['F'+str(row)] = round(df.loc[(df['Geography'] == 'OTSA') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'CV'].iat[0], 2)
        ws['G'+str(row)] = round(df.loc[(df['Geography'] == 'OTSA') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'MeanPercDiff'].iat[0], 2)
        ws['H'+str(row)] = df.loc[(df['Geography'] == 'OTSA') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'NumberGreater10Perc'].iat[0]
        row += 1

# Table 7i
row = 908
for s in smallsizecats_plusall:
    row += 1
    for r in raceincombcats:
        ws['B'+str(row)] = df.loc[(df['Geography'] == 'ANVSA') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'NumCells'].iat[0]
        ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'ANVSA') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'MeanAbsDiff'].iat[0], 2)
        ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'ANVSA') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'RMSE'].iat[0], 2)
        ws['E'+str(row)] = round(df.loc[(df['Geography'] == 'ANVSA') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'MeanAbsPercDiff'].iat[0], 2)
        ws['F'+str(row)] = round(df.loc[(df['Geography'] == 'ANVSA') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'CV'].iat[0], 2)
        ws['G'+str(row)] = round(df.loc[(df['Geography'] == 'ANVSA') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'MeanPercDiff'].iat[0], 2)
        ws['H'+str(row)] = df.loc[(df['Geography'] == 'ANVSA') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'NumberGreater10Perc'].iat[0]
        row += 1


# Table 8a
row = 941
for ho in ['Hispanic', 'Non-Hispanic']:
    row += 1
    for r in racealonecats:
        ws['B'+str(row)] = df.loc[(df['Geography'] == 'State') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{ho} {r}"), 'NumCells'].iat[0]
        ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'State') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{ho} {r}"), 'MeanAbsDiff'].iat[0], 2)
        ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'State') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{ho} {r}"), 'RMSE'].iat[0], 2)
        ws['E'+str(row)] = round(df.loc[(df['Geography'] == 'State') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{ho} {r}"), 'MeanAbsPercDiff'].iat[0], 2)
        ws['F'+str(row)] = round(df.loc[(df['Geography'] == 'State') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{ho} {r}"), 'CV'].iat[0], 2)
        ws['G'+str(row)] = round(df.loc[(df['Geography'] == 'State') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{ho} {r}"), 'MeanPercDiff'].iat[0], 2)
        ws['H'+str(row)] = df.loc[(df['Geography'] == 'State') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{ho} {r}"), 'NumberGreater10Perc'].iat[0]
        row += 1

# Table 8b
row = 962
for s in smallsizecats_plusall:
    row += 1
    for ho in ['Hispanic', 'Non-Hispanic']:
        row += 1
        for r in racealonecats:
            ws['B'+str(row)] = df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{ho} {r}"), 'NumCells'].iat[0]
            ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{ho} {r}"), 'MeanAbsDiff'].iat[0], 2)
            ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{ho} {r}"), 'RMSE'].iat[0], 2)
            ws['E'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{ho} {r}"), 'MeanAbsPercDiff'].iat[0], 2)
            ws['F'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{ho} {r}"), 'CV'].iat[0], 2)
            ws['G'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{ho} {r}"), 'MeanPercDiff'].iat[0], 2)
            ws['H'+str(row)] = df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{ho} {r}"), 'NumberGreater10Perc'].iat[0]
            row += 1

# Table 8c
row = 1035
for s in smallsizecats_plusall:
    row += 1
    for ho in ['Hispanic', 'Non-Hispanic']:
        row += 1
        for r in racealonecats:
            ws['B'+str(row)] = df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{ho} {r}"), 'NumCells'].iat[0]
            ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{ho} {r}"), 'MeanAbsDiff'].iat[0], 2)
            ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{ho} {r}"), 'RMSE'].iat[0], 2)
            ws['E'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{ho} {r}"), 'MeanAbsPercDiff'].iat[0], 2)
            ws['F'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{ho} {r}"), 'CV'].iat[0], 2)
            ws['G'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{ho} {r}"), 'MeanPercDiff'].iat[0], 2)
            ws['H'+str(row)] = df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{ho} {r}"), 'NumberGreater10Perc'].iat[0]
            row += 1

# Table 8d
row = 1108
for s in smallsizecats_plusall:
    row += 1
    for ho in ['Hispanic', 'Non-Hispanic']:
        row += 1
        for r in racealonecats:
            ws['B'+str(row)] = df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{ho} {r}"), 'NumCells'].iat[0]
            ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{ho} {r}"), 'MeanAbsDiff'].iat[0], 2)
            ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{ho} {r}"), 'RMSE'].iat[0], 2)
            ws['E'+str(row)] = df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{ho} {r}"), 'NumberGreater10Perc'].iat[0]
            row += 1

# Table 9a
row = 1181
for ho in ['Hispanic', 'Non-Hispanic']:
    row += 1
    for r in raceincombcats:
        ws['B'+str(row)] = df.loc[(df['Geography'] == 'State') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{ho} {r}"), 'NumCells'].iat[0]
        ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'State') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{ho} {r}"), 'MeanAbsDiff'].iat[0], 2)
        ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'State') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{ho} {r}"), 'RMSE'].iat[0], 2)
        ws['E'+str(row)] = round(df.loc[(df['Geography'] == 'State') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{ho} {r}"), 'MeanAbsPercDiff'].iat[0], 2)
        ws['F'+str(row)] = round(df.loc[(df['Geography'] == 'State') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{ho} {r}"), 'CV'].iat[0], 2)
        ws['G'+str(row)] = round(df.loc[(df['Geography'] == 'State') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{ho} {r}"), 'MeanPercDiff'].iat[0], 2)
        ws['H'+str(row)] = df.loc[(df['Geography'] == 'State') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{ho} {r}"), 'NumberGreater10Perc'].iat[0]
        row += 1

# Table 9b
row = 1200
for s in smallsizecats_plusall:
    row += 1
    for ho in ['Hispanic', 'Non-Hispanic']:
        row += 1
        for r in raceincombcats:
            ws['B'+str(row)] = df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{ho} {r}"), 'NumCells'].iat[0]
            ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{ho} {r}"), 'MeanAbsDiff'].iat[0], 2)
            ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{ho} {r}"), 'RMSE'].iat[0], 2)
            ws['E'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{ho} {r}"), 'MeanAbsPercDiff'].iat[0], 2)
            ws['F'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{ho} {r}"), 'CV'].iat[0], 2)
            ws['G'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{ho} {r}"), 'MeanPercDiff'].iat[0], 2)
            ws['H'+str(row)] = df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{ho} {r}"), 'NumberGreater10Perc'].iat[0]
            row += 1

# Table 9c
row = 1265
for s in smallsizecats_plusall:
    row += 1
    for ho in ['Hispanic', 'Non-Hispanic']:
        row += 1
        for r in raceincombcats:
            ws['B'+str(row)] = df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{ho} {r}"), 'NumCells'].iat[0]
            ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{ho} {r}"), 'MeanAbsDiff'].iat[0], 2)
            ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{ho} {r}"), 'RMSE'].iat[0], 2)
            ws['E'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{ho} {r}"), 'MeanAbsPercDiff'].iat[0], 2)
            ws['F'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{ho} {r}"), 'CV'].iat[0], 2)
            ws['G'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{ho} {r}"), 'MeanPercDiff'].iat[0], 2)
            ws['H'+str(row)] = df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{ho} {r}"), 'NumberGreater10Perc'].iat[0]
            row += 1

# Table 9d
row = 1330
for s in smallsizecats_plusall:
    row += 1
    for ho in ['Hispanic', 'Non-Hispanic']:
        row += 1
        for r in raceincombcats:
            ws['B'+str(row)] = df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{ho} {r}"), 'NumCells'].iat[0]
            ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{ho} {r}"), 'MeanAbsDiff'].iat[0], 2)
            ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{ho} {r}"), 'RMSE'].iat[0], 2)
            ws['E'+str(row)] = df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{ho} {r}"), 'NumberGreater10Perc'].iat[0]
            row += 1

# Table 10a
row = 1395
for r in numracecats:
    ws['B'+str(row)] = df.loc[(df['Geography'] == 'State') & (df['Size_Category'] == "All") & (df['Characteristic'] == r), 'NumCells'].iat[0]
    ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'State') & (df['Size_Category'] == "All") & (df['Characteristic'] == r), 'MeanAbsDiff'].iat[0], 2)
    ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'State') & (df['Size_Category'] == "All") & (df['Characteristic'] == r), 'RMSE'].iat[0], 2)
    ws['E'+str(row)] = round(df.loc[(df['Geography'] == 'State') & (df['Size_Category'] == "All") & (df['Characteristic'] == r), 'MeanAbsPercDiff'].iat[0], 2)
    ws['F'+str(row)] = round(df.loc[(df['Geography'] == 'State') & (df['Size_Category'] == "All") & (df['Characteristic'] == r), 'CV'].iat[0], 2)
    ws['G'+str(row)] = round(df.loc[(df['Geography'] == 'State') & (df['Size_Category'] == "All") & (df['Characteristic'] == r), 'MeanPercDiff'].iat[0], 2)
    ws['H'+str(row)] = df.loc[(df['Geography'] == 'State') & (df['Size_Category'] == "All") & (df['Characteristic'] == r), 'NumberGreater10Perc'].iat[0]
    row += 1

# Table 10b
row = 1406
for s in smallsizecats_plusall:
    row += 1
    for r in numracecats:
        ws['B'+str(row)] = df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'NumCells'].iat[0]
        ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'MeanAbsDiff'].iat[0], 2)
        ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'RMSE'].iat[0], 2)
        ws['E'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'MeanAbsPercDiff'].iat[0], 2)
        ws['F'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'CV'].iat[0], 2)
        ws['G'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'MeanPercDiff'].iat[0], 2)
        ws['H'+str(row)] = df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'NumberGreater10Perc'].iat[0]
        row += 1

# Table 10c
row = 1439
for s in smallsizecats_plusall:
    row += 1
    for r in numracecats:
        ws['B'+str(row)] = df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'NumCells'].iat[0]
        ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'MeanAbsDiff'].iat[0], 2)
        ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'RMSE'].iat[0], 2)
        ws['E'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'MeanAbsPercDiff'].iat[0], 2)
        ws['F'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'CV'].iat[0], 2)
        ws['G'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'MeanPercDiff'].iat[0], 2)
        ws['H'+str(row)] = df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'NumberGreater10Perc'].iat[0]
        row += 1

# Table 10d
if tractnumracessizesready:
    row = 1472
    for s in smallsizecats_plusall:
        row += 1
        for r in numracecats:
            ws['B'+str(row)] = df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'NumCells'].iat[0]
            ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'MeanAbsDiff'].iat[0], 2)
            ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'RMSE'].iat[0], 2)
            ws['E'+str(row)] = df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == s) & (df['Characteristic'] == r), 'NumberGreater10Perc'].iat[0]
            row += 1

# Table 11a
row = 1505
for ho in ['Hispanic', 'Non-Hispanic']:
    row += 1
    for r in numracecats:
        ws['B'+str(row)] = df.loc[(df['Geography'] == 'State') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{ho} {r}"), 'NumCells'].iat[0]
        ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'State') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{ho} {r}"), 'MeanAbsDiff'].iat[0], 2)
        ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'State') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{ho} {r}"), 'RMSE'].iat[0], 2)
        ws['E'+str(row)] = round(df.loc[(df['Geography'] == 'State') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{ho} {r}"), 'MeanAbsPercDiff'].iat[0], 2)
        ws['F'+str(row)] = round(df.loc[(df['Geography'] == 'State') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{ho} {r}"), 'CV'].iat[0], 2)
        ws['G'+str(row)] = round(df.loc[(df['Geography'] == 'State') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{ho} {r}"), 'MeanPercDiff'].iat[0], 2)
        ws['H'+str(row)] = df.loc[(df['Geography'] == 'State') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{ho} {r}"), 'NumberGreater10Perc'].iat[0]
        row += 1

# Table 11b
row = 1524
for s in smallsizecats_plusall:
    row += 1
    for ho in ['Hispanic', 'Non-Hispanic']:
        row += 1
        for r in numracecats:
            ws['B'+str(row)] = df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{ho} {r}"), 'NumCells'].iat[0]
            ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{ho} {r}"), 'MeanAbsDiff'].iat[0], 2)
            ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{ho} {r}"), 'RMSE'].iat[0], 2)
            ws['E'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{ho} {r}"), 'MeanAbsPercDiff'].iat[0], 2)
            ws['F'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{ho} {r}"), 'CV'].iat[0], 2)
            ws['G'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{ho} {r}"), 'MeanPercDiff'].iat[0], 2)
            ws['H'+str(row)] = df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{ho} {r}"), 'NumberGreater10Perc'].iat[0]
            row += 1

# Table 11c
row = 1589
for s in smallsizecats_plusall:
    row += 1
    for ho in ['Hispanic', 'Non-Hispanic']:
        row += 1
        for r in numracecats:
            ws['B'+str(row)] = df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{ho} {r}"), 'NumCells'].iat[0]
            ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{ho} {r}"), 'MeanAbsDiff'].iat[0], 2)
            ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{ho} {r}"), 'RMSE'].iat[0], 2)
            ws['E'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{ho} {r}"), 'MeanAbsPercDiff'].iat[0], 2)
            ws['F'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{ho} {r}"), 'CV'].iat[0], 2)
            ws['G'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{ho} {r}"), 'MeanPercDiff'].iat[0], 2)
            ws['H'+str(row)] = df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{ho} {r}"), 'NumberGreater10Perc'].iat[0]
            row += 1

# Table 11d
row = 1654
for s in smallsizecats_plusall:
    row += 1
    for ho in ['Hispanic', 'Non-Hispanic']:
        row += 1
        for r in numracecats:
            ws['B'+str(row)] = df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{ho} {r}"), 'NumCells'].iat[0]
            ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{ho} {r}"), 'MeanAbsDiff'].iat[0], 2)
            ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{ho} {r}"), 'RMSE'].iat[0], 2)
            ws['E'+str(row)] = df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{ho} {r}"), 'NumberGreater10Perc'].iat[0]
            row += 1

# Table 12a
row = 1719
for ho in ['Hispanic', 'Non-Hispanic']:
    row += 1
    for r in racealonecats:
        ws['B'+str(row)] = df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == f"{ho} {r} Aged 18+"), 'NumCells'].iat[0]
        ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == f"{ho} {r} Aged 18+"), 'MeanAbsDiff'].iat[0], 2)
        ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == f"{ho} {r} Aged 18+"), 'RMSE'].iat[0], 2)
        ws['E'+str(row)] = df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == f"{ho} {r} Aged 18+"), 'NumberGreater10Perc'].iat[0]
        row += 1

# Table 12b
row = 1740
for ho in ['Hispanic', 'Non-Hispanic']:
    row += 1
    for r in racealonecats:
        ws['B'+str(row)] = df.loc[(df['Geography'] == 'Block Group') & (df['Size_Category'] == 'All') & (df['Characteristic'] == f"{ho} {r} Aged 18+"), 'NumCells'].iat[0]
        ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'Block Group') & (df['Size_Category'] == 'All') & (df['Characteristic'] == f"{ho} {r} Aged 18+"), 'MeanAbsDiff'].iat[0], 2)
        ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'Block Group') & (df['Size_Category'] == 'All') & (df['Characteristic'] == f"{ho} {r} Aged 18+"), 'RMSE'].iat[0], 2)
        ws['E'+str(row)] = df.loc[(df['Geography'] == 'Block Group') & (df['Size_Category'] == 'All') & (df['Characteristic'] == f"{ho} {r} Aged 18+"), 'NumberGreater10Perc'].iat[0]
        row += 1

# Table 13a
row = 1761
for ho in ['Hispanic', 'Non-Hispanic']:
    row += 1
    for r in raceincombcats:
        ws['B'+str(row)] = df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == f"{ho} {r} Aged 18+"), 'NumCells'].iat[0]
        ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == f"{ho} {r} Aged 18+"), 'MeanAbsDiff'].iat[0], 2)
        ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == f"{ho} {r} Aged 18+"), 'RMSE'].iat[0], 2)
        ws['E'+str(row)] = df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == f"{ho} {r} Aged 18+"), 'NumberGreater10Perc'].iat[0]
        row += 1

# Table 13b
row = 1780
for ho in ['Hispanic', 'Non-Hispanic']:
    row += 1
    for r in raceincombcats:
        ws['B'+str(row)] = df.loc[(df['Geography'] == 'Block Group') & (df['Size_Category'] == 'All') & (df['Characteristic'] == f"{ho} {r} Aged 18+"), 'NumCells'].iat[0]
        ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'Block Group') & (df['Size_Category'] == 'All') & (df['Characteristic'] == f"{ho} {r} Aged 18+"), 'MeanAbsDiff'].iat[0], 2)
        ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'Block Group') & (df['Size_Category'] == 'All') & (df['Characteristic'] == f"{ho} {r} Aged 18+"), 'RMSE'].iat[0], 2)
        ws['E'+str(row)] = df.loc[(df['Geography'] == 'Block Group') & (df['Size_Category'] == 'All') & (df['Characteristic'] == f"{ho} {r} Aged 18+"), 'NumberGreater10Perc'].iat[0]
        row += 1

# Table 14a
row = 1799
for ho in ['Hispanic', 'Non-Hispanic']:
    row += 1
    for r in numracecats:
        ws['B'+str(row)] = df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == f"{ho} {r} Aged 18+"), 'NumCells'].iat[0]
        ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == f"{ho} {r} Aged 18+"), 'MeanAbsDiff'].iat[0], 2)
        ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == f"{ho} {r} Aged 18+"), 'RMSE'].iat[0], 2)
        ws['E'+str(row)] = df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == f"{ho} {r} Aged 18+"), 'NumberGreater10Perc'].iat[0]
        row += 1

# Table 14b
row = 1818
for ho in ['Hispanic', 'Non-Hispanic']:
    row += 1
    for r in numracecats:
        ws['B'+str(row)] = df.loc[(df['Geography'] == 'Block Group') & (df['Size_Category'] == 'All') & (df['Characteristic'] == f"{ho} {r} Aged 18+"), 'NumCells'].iat[0]
        ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'Block Group') & (df['Size_Category'] == 'All') & (df['Characteristic'] == f"{ho} {r} Aged 18+"), 'MeanAbsDiff'].iat[0], 2)
        ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'Block Group') & (df['Size_Category'] == 'All') & (df['Characteristic'] == f"{ho} {r} Aged 18+"), 'RMSE'].iat[0], 2)
        ws['E'+str(row)] = df.loc[(df['Geography'] == 'Block Group') & (df['Size_Category'] == 'All') & (df['Characteristic'] == f"{ho} {r} Aged 18+"), 'NumberGreater10Perc'].iat[0]
        row += 1

# Table 15a
row = 1837
for s in ["All","[0.0, 1000.0)"]:
    row += 1
    for sex in ["","Male ","Female "]:
        row += 1
        for a in ["Age [0, 18)","Age [18, 65)", "Age [65, 116)"]:
            ws['B'+str(row)] = df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{sex}{a}"), 'NumCells'].iat[0]
            ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{sex}{a}"), 'MeanAbsDiff'].iat[0], 2)
            ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{sex}{a}"), 'RMSE'].iat[0], 2)
            ws['E'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{sex}{a}"), 'MeanAbsPercDiff'].iat[0], 2)
            ws['F'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{sex}{a}"), 'CV'].iat[0], 2)
            ws['G'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{sex}{a}"), 'MeanPercDiff'].iat[0], 2)
            ws['H'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{sex}{a}"), 'NumberGreater10Perc'].iat[0], 2)
            row += 1

# Table 15b
row = 1868
for s in ["All","[0.0, 500.0)"]:
    row += 1
    for sex in ["","Male ","Female "]:
        row += 1
        for a in ["Age [0, 18)","Age [18, 65)", "Age [65, 116)"]:
            ws['B'+str(row)] = df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{sex}{a}"), 'NumCells'].iat[0]
            ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{sex}{a}"), 'MeanAbsDiff'].iat[0], 2)
            ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{sex}{a}"), 'RMSE'].iat[0], 2)
            ws['E'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{sex}{a}"), 'MeanAbsPercDiff'].iat[0], 2)
            ws['F'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{sex}{a}"), 'CV'].iat[0], 2)
            ws['G'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{sex}{a}"), 'MeanPercDiff'].iat[0], 2)
            ws['H'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == f"{sex}{a}"), 'NumberGreater10Perc'].iat[0], 2)
            row += 1

# Table 15c
row = 1899
for sex in ["","Male ","Female "]:
    row += 1
    for a in ["Age [0, 18)","Age [18, 65)", "Age [65, 116)"]:
        ws['B'+str(row)] = df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{sex}{a}"), 'NumCells'].iat[0]
        ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{sex}{a}"), 'MeanAbsDiff'].iat[0], 2)
        ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{sex}{a}"), 'RMSE'].iat[0], 2)
        ws['E'+str(row)] = round(df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{sex}{a}"), 'NumberGreater10Perc'].iat[0], 2)
        row += 1

# Table 16a
row = 1916
for sex in ["","Male ","Female "]:
    row += 1
    for a in ["Age [0, 5)", "Age [5, 10)", "Age [10, 15)", "Age [15, 20)", "Age [20, 25)", "Age [25, 30)", "Age [30, 35)", "Age [35, 40)", "Age [40, 45)", "Age [45, 50)", "Age [50, 55)", "Age [55, 60)", "Age [60, 65)", "Age [65, 70)", "Age [70, 75)", "Age [75, 80)", "Age [80, 85)", "Age [85, 90)", "Age [90, 95)", "Age [95, 100)", "Age [100, 105)", "Age [105, 110)", "Age [110, 116)"]:
        ws['B'+str(row)] = df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{sex}{a}"), 'NumCells'].iat[0]
        ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{sex}{a}"), 'MeanAbsDiff'].iat[0], 2)
        ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{sex}{a}"), 'RMSE'].iat[0], 2)
        ws['E'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{sex}{a}"), 'MeanAbsPercDiff'].iat[0], 2)
        ws['F'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{sex}{a}"), 'CV'].iat[0], 2)
        ws['G'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{sex}{a}"), 'MeanPercDiff'].iat[0], 2)
        ws['H'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{sex}{a}"), 'NumberGreater10Perc'].iat[0], 2)
        row += 1

# Table 16b
row = 1993
for sex in ["","Male ","Female "]:
    row += 1
    for a in ["Age [0, 5)", "Age [5, 10)", "Age [10, 15)", "Age [15, 20)", "Age [20, 25)", "Age [25, 30)", "Age [30, 35)", "Age [35, 40)", "Age [40, 45)", "Age [45, 50)", "Age [50, 55)", "Age [55, 60)", "Age [60, 65)", "Age [65, 70)", "Age [70, 75)", "Age [75, 80)", "Age [80, 85)", "Age [85, 90)", "Age [90, 95)", "Age [95, 100)", "Age [100, 105)", "Age [105, 110)", "Age [110, 116)"]:
        ws['B'+str(row)] = df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{sex}{a}"), 'NumCells'].iat[0]
        ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{sex}{a}"), 'MeanAbsDiff'].iat[0], 2)
        ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{sex}{a}"), 'RMSE'].iat[0], 2)
        ws['E'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{sex}{a}"), 'MeanAbsPercDiff'].iat[0], 2)
        ws['F'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{sex}{a}"), 'CV'].iat[0], 2)
        ws['G'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{sex}{a}"), 'MeanPercDiff'].iat[0], 2)
        ws['H'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{sex}{a}"), 'NumberGreater10Perc'].iat[0], 2)
        row += 1

# Table 16c
row = 2070
for sex in ["","Male ","Female "]:
    row += 1
    for a in ["Age [0, 5)", "Age [5, 10)", "Age [10, 15)", "Age [15, 20)", "Age [20, 25)", "Age [25, 30)", "Age [30, 35)", "Age [35, 40)", "Age [40, 45)", "Age [45, 50)", "Age [50, 55)", "Age [55, 60)", "Age [60, 65)", "Age [65, 70)", "Age [70, 75)", "Age [75, 80)", "Age [80, 85)", "Age [85, 90)", "Age [90, 95)", "Age [95, 100)", "Age [100, 105)", "Age [105, 110)", "Age [110, 116)"]:
        ws['B'+str(row)] = df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{sex}{a}"), 'NumCells'].iat[0]
        ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{sex}{a}"), 'MeanAbsDiff'].iat[0], 2)
        ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{sex}{a}"), 'RMSE'].iat[0], 2)
        ws['E'+str(row)] = round(df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{sex}{a}"), 'NumberGreater10Perc'].iat[0], 2)
        row += 1

if prmetricsready:
    # Table 16d
    row = 2147
    for sex in ["","Male ","Female "]:
        row += 1
        for a in ["Age [0, 5)", "Age [5, 10)", "Age [10, 15)", "Age [15, 20)", "Age [20, 25)", "Age [25, 30)", "Age [30, 35)", "Age [35, 40)", "Age [40, 45)", "Age [45, 50)", "Age [50, 55)", "Age [55, 60)", "Age [60, 65)", "Age [65, 70)", "Age [70, 75)", "Age [75, 80)", "Age [80, 85)", "Age [85, 90)", "Age [90, 95)", "Age [95, 100)", "Age [100, 105)", "Age [105, 110)", "Age [110, 116)"]:
            ws['B'+str(row)] = df.loc[(df['Geography'] == 'PR County/Municipio') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{sex}{a}"), 'NumCells'].iat[0]
            ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'PR County/Municipio') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{sex}{a}"), 'MeanAbsDiff'].iat[0], 2)
            ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'PR County/Municipio') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{sex}{a}"), 'RMSE'].iat[0], 2)
            ws['E'+str(row)] = round(df.loc[(df['Geography'] == 'PR County/Municipio') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{sex}{a}"), 'MeanAbsPercDiff'].iat[0], 2)
            ws['F'+str(row)] = round(df.loc[(df['Geography'] == 'PR County/Municipio') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{sex}{a}"), 'CV'].iat[0], 2)
            ws['G'+str(row)] = round(df.loc[(df['Geography'] == 'PR County/Municipio') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{sex}{a}"), 'MeanPercDiff'].iat[0], 2)
            ws['H'+str(row)] = round(df.loc[(df['Geography'] == 'PR County/Municipio') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{sex}{a}"), 'NumberGreater10Perc'].iat[0], 2)
            row += 1

    # Table 16e
    row = 2224
    for sex in ["","Male ","Female "]:
        row += 1
        for a in ["Age [0, 5)", "Age [5, 10)", "Age [10, 15)", "Age [15, 20)", "Age [20, 25)", "Age [25, 30)", "Age [30, 35)", "Age [35, 40)", "Age [40, 45)", "Age [45, 50)", "Age [50, 55)", "Age [55, 60)", "Age [60, 65)", "Age [65, 70)", "Age [70, 75)", "Age [75, 80)", "Age [80, 85)", "Age [85, 90)", "Age [90, 95)", "Age [95, 100)", "Age [100, 105)", "Age [105, 110)", "Age [110, 116)"]:
            ws['B'+str(row)] = df.loc[(df['Geography'] == 'PR Tract') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{sex}{a}"), 'NumCells'].iat[0]
            ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'PR Tract') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{sex}{a}"), 'MeanAbsDiff'].iat[0], 2)
            ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'PR Tract') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{sex}{a}"), 'RMSE'].iat[0], 2)
            ws['E'+str(row)] = round(df.loc[(df['Geography'] == 'PR Tract') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{sex}{a}"), 'NumberGreater10Perc'].iat[0], 2)
            row += 1

# Table 16f
row = 2301
for sex in ["","Male ","Female "]:
    row += 1
    for a in ["Age [0, 5)", "Age [5, 10)", "Age [10, 15)", "Age [15, 20)", "Age [20, 25)", "Age [25, 30)", "Age [30, 35)", "Age [35, 40)", "Age [40, 45)", "Age [45, 50)", "Age [50, 55)", "Age [55, 60)", "Age [60, 65)", "Age [65, 70)", "Age [70, 75)", "Age [75, 80)", "Age [80, 85)", "Age [85, 90)", "Age [90, 95)", "Age [95, 100)", "Age [100, 105)", "Age [105, 110)", "Age [110, 116)"]:
        ws['B'+str(row)] = df.loc[(df['Geography'] == 'Fed AIR') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{sex}{a}"), 'NumCells'].iat[0]
        ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'Fed AIR') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{sex}{a}"), 'MeanAbsDiff'].iat[0], 2)
        ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'Fed AIR') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{sex}{a}"), 'RMSE'].iat[0], 2)
        ws['E'+str(row)] = round(df.loc[(df['Geography'] == 'Fed AIR') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{sex}{a}"), 'MeanAbsPercDiff'].iat[0], 2)
        ws['F'+str(row)] = round(df.loc[(df['Geography'] == 'Fed AIR') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{sex}{a}"), 'CV'].iat[0], 2)
        ws['G'+str(row)] = round(df.loc[(df['Geography'] == 'Fed AIR') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{sex}{a}"), 'MeanPercDiff'].iat[0], 2)
        ws['H'+str(row)] = round(df.loc[(df['Geography'] == 'Fed AIR') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{sex}{a}"), 'NumberGreater10Perc'].iat[0], 2)
        row += 1

# Table 16g
row = 2378
for sex in ["","Male ","Female "]:
    row += 1
    for a in ["Age [0, 5)", "Age [5, 10)", "Age [10, 15)", "Age [15, 20)", "Age [20, 25)", "Age [25, 30)", "Age [30, 35)", "Age [35, 40)", "Age [40, 45)", "Age [45, 50)", "Age [50, 55)", "Age [55, 60)", "Age [60, 65)", "Age [65, 70)", "Age [70, 75)", "Age [75, 80)", "Age [80, 85)", "Age [85, 90)", "Age [90, 95)", "Age [95, 100)", "Age [100, 105)", "Age [105, 110)", "Age [110, 116)"]:
        ws['B'+str(row)] = df.loc[(df['Geography'] == 'OTSA') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{sex}{a}"), 'NumCells'].iat[0]
        ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'OTSA') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{sex}{a}"), 'MeanAbsDiff'].iat[0], 2)
        ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'OTSA') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{sex}{a}"), 'RMSE'].iat[0], 2)
        ws['E'+str(row)] = round(df.loc[(df['Geography'] == 'OTSA') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{sex}{a}"), 'MeanAbsPercDiff'].iat[0], 2)
        ws['F'+str(row)] = round(df.loc[(df['Geography'] == 'OTSA') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{sex}{a}"), 'CV'].iat[0], 2)
        ws['G'+str(row)] = round(df.loc[(df['Geography'] == 'OTSA') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{sex}{a}"), 'MeanPercDiff'].iat[0], 2)
        ws['H'+str(row)] = round(df.loc[(df['Geography'] == 'OTSA') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{sex}{a}"), 'NumberGreater10Perc'].iat[0], 2)
        row += 1

# Table 16h
row = 2455
for sex in ["","Male ","Female "]:
    row += 1
    for a in ["Age [0, 5)", "Age [5, 10)", "Age [10, 15)", "Age [15, 20)", "Age [20, 25)", "Age [25, 30)", "Age [30, 35)", "Age [35, 40)", "Age [40, 45)", "Age [45, 50)", "Age [50, 55)", "Age [55, 60)", "Age [60, 65)", "Age [65, 70)", "Age [70, 75)", "Age [75, 80)", "Age [80, 85)", "Age [85, 90)", "Age [90, 95)", "Age [95, 100)", "Age [100, 105)", "Age [105, 110)", "Age [110, 116)"]:
        ws['B'+str(row)] = df.loc[(df['Geography'] == 'ANVSA') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{sex}{a}"), 'NumCells'].iat[0]
        ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'ANVSA') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{sex}{a}"), 'MeanAbsDiff'].iat[0], 2)
        ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'ANVSA') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{sex}{a}"), 'RMSE'].iat[0], 2)
        ws['E'+str(row)] = round(df.loc[(df['Geography'] == 'ANVSA') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{sex}{a}"), 'MeanAbsPercDiff'].iat[0], 2)
        ws['F'+str(row)] = round(df.loc[(df['Geography'] == 'ANVSA') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{sex}{a}"), 'CV'].iat[0], 2)
        ws['G'+str(row)] = round(df.loc[(df['Geography'] == 'ANVSA') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{sex}{a}"), 'MeanPercDiff'].iat[0], 2)
        ws['H'+str(row)] = round(df.loc[(df['Geography'] == 'ANVSA') & (df['Size_Category'] == "All") & (df['Characteristic'] == f"{sex}{a}"), 'NumberGreater10Perc'].iat[0], 2)
        row += 1

# Table 17
row = 2532
for s in sizecats_plusall:
    ws['B'+str(row)] = df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == "Average Absolute Change in Median Age"), 'NumCells'].iat[0]
    ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == "Average Absolute Change in Median Age"), 'AvgAbsDiffMedAge'].iat[0], 2)
    ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == "Average Absolute Change in Sex Ratio"), 'AvgAbsDiffSexRatio'].iat[0], 2)
    row += 1

# Table 18a
row = 2544
row += 1
for c in ["GQ INST","GQ Correctional", "GQ Juvenile", "GQ Nursing", "GQ OtherInst", "GQ NONINST", "GQ College", "GQ Military", "GQ OtherNoninst"]:
    ws['B'+str(row)] = df.loc[(df['Geography'] == 'State') & (df['Size_Category'] == 'All') & (df['Characteristic'] == c), 'NumCells'].iat[0]
    ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'State') & (df['Size_Category'] == 'All') & (df['Characteristic'] == c), 'MeanAbsDiff'].iat[0], 2)
    ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'State') & (df['Size_Category'] == 'All') & (df['Characteristic'] == c), 'RMSE'].iat[0], 2)
    ws['E'+str(row)] = round(df.loc[(df['Geography'] == 'State') & (df['Size_Category'] == 'All') & (df['Characteristic'] == c), 'MeanAbsPercDiff'].iat[0], 2)
    ws['F'+str(row)] = round(df.loc[(df['Geography'] == 'State') & (df['Size_Category'] == 'All') & (df['Characteristic'] == c), 'CV'].iat[0], 2)
    ws['G'+str(row)] = round(df.loc[(df['Geography'] == 'State') & (df['Size_Category'] == 'All') & (df['Characteristic'] == c), 'MeanPercDiff'].iat[0], 2)
    ws['H'+str(row)] = round(df.loc[(df['Geography'] == 'State') & (df['Size_Category'] == 'All') & (df['Characteristic'] == c), 'NumberGreater10Perc'].iat[0], 2)
    row += 1

# Table 18b
row = 2559
for s in sizecats_plusall:
    row += 1
    for c in ["GQ INST","GQ Correctional", "GQ Juvenile", "GQ Nursing", "GQ OtherInst", "GQ NONINST", "GQ College", "GQ Military", "GQ OtherNoninst"]:
        ws['B'+str(row)] = df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == c), 'NumCells'].iat[0]
        ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == c), 'MeanAbsDiff'].iat[0], 2)
        ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == c), 'RMSE'].iat[0], 2)
        ws['E'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == c), 'MeanAbsPercDiff'].iat[0], 2)
        ws['F'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == c), 'CV'].iat[0], 2)
        ws['G'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == c), 'MeanPercDiff'].iat[0], 2)
        ws['H'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == c), 'NumberGreater10Perc'].iat[0], 2)
        row += 1

# Table 18c
row = 2634
for s in incplacesizecats_plusall:
    row += 1
    for c in ["GQ INST","GQ Correctional", "GQ Juvenile", "GQ Nursing", "GQ OtherInst", "GQ NONINST", "GQ College", "GQ Military", "GQ OtherNoninst"]:
        ws['B'+str(row)] = df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == c), 'NumCells'].iat[0]
        ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == c), 'MeanAbsDiff'].iat[0], 2)
        ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == c), 'RMSE'].iat[0], 2)
        ws['E'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == c), 'MeanAbsPercDiff'].iat[0], 2)
        ws['F'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == c), 'CV'].iat[0], 2)
        ws['G'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == c), 'MeanPercDiff'].iat[0], 2)
        ws['H'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == c), 'NumberGreater10Perc'].iat[0], 2)
        row += 1

# Table 18d
row = 2719
for c in ["GQ INST","GQ Correctional", "GQ Juvenile", "GQ Nursing", "GQ OtherInst", "GQ NONINST", "GQ College", "GQ Military", "GQ OtherNoninst"]:
    ws['B'+str(row)] = df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == c), 'NumCells'].iat[0]
    ws['C'+str(row)] = round(df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == c), 'MeanAbsDiff'].iat[0], 2)
    ws['D'+str(row)] = round(df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == c), 'RMSE'].iat[0], 2)
    ws['E'+str(row)] = round(df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == c), 'NumberGreater10Perc'].iat[0], 2)
    row += 1


ws1 = wb["Updated Use Cases"]

# Table 1
row = 8
ws1['B'+str(row)] = df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Aged 75 and Over"), 'NumCells'].iat[0]
ws1['C'+str(row)] = round(df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Aged 75 and Over"), 'MeanAbsDiff'].iat[0], 2)
ws1['D'+str(row)] = round(df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Aged 75 and Over"), 'RMSE'].iat[0], 2)
ws1['E'+str(row)] = df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Aged 75 and Over"), 'NumberBtw2Perc5Perc'].iat[0]
ws1['F'+str(row)] = df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Aged 75 and Over"), 'NumberGreater5Perc'].iat[0]

# Table 2
statelist = ["AL", "AK", "AZ", "AR", "CA", "CO", "CT", "DE", "DC", "FL", "GA", "HI", "ID", "IL", "IN", "IA", "KS", "KY", "LA", "ME", "MD", "MA", "MI", "MN", "MS", "MO", "MT", "NE", "NV", "NH", "NJ", "NM", "NY", "NC", "ND", "OH", "OK", "OR", "PA", "RI", "SC", "SD", "TN", "TX", "UT", "VT", "VA", "WA", "WV", "WI", "WY"]
row = 15
for s in statelist:
    ws1['B'+str(row)] = df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "TAES Total Population {}".format(s)), "NumCells"].iat[0]
    ws1['C'+str(row)] = df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "TAES Total Population {}".format(s)), "TAES"].iat[0]
    row += 1

# Table 3
statelist = ["AL", "AK", "AZ", "AR", "CA", "CO", "CT", "DE", "DC", "FL", "GA", "HI", "ID", "IL", "IN", "IA", "KS", "KY", "LA", "ME", "MD", "MA", "MI", "MN", "MS", "MO", "MT", "NE", "NV", "NH", "NJ", "NM", "NY", "NC", "ND", "OH", "OK", "OR", "PA", "RI", "SC", "SD", "TN", "TX", "UT", "VT", "VA", "WA", "WV", "WI", "WY"]
row = 71
for s in statelist:
    ws1['B'+str(row)] = df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "TAES Total Population {}".format(s)), "NumCells"].iat[0]
    ws1['C'+str(row)] = df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "TAES Total Population {}".format(s)), "TAES"].iat[0]
    row += 1

# Table 4
statelist = ["CT", "ME", "MA", "MI", "MN", "NH", "NJ", "NY", "PA", "RI", "VT", "WI"]
row = 127
for s in statelist:
    ws1['B'+str(row)] = df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "TAES Total Population {}".format(s)), "NumCells"].iat[0]
    ws1['C'+str(row)] = df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "TAES Total Population {}".format(s)), "TAES"].iat[0]
    row += 1

# Table 5a
row = 145
for s in sizecats_u18_plusall:
    row += 1
    for c in range(0,18):
        ws1['B'+str(row)] = df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == f"Age {c}"), 'NumCells'].iat[0]
        ws1['C'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == f"Age {c}"), 'MeanAbsDiff'].iat[0], 2)
        ws1['D'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == f"Age {c}"), 'RMSE'].iat[0], 2)
        ws1['E'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == f"Age {c}"), 'MeanAbsPercDiff'].iat[0], 2)
        ws1['F'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == f"Age {c}"), 'CV'].iat[0], 2)
        ws1['G'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == f"Age {c}"), 'MeanPercDiff'].iat[0], 2)
        ws1['H'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == f"Age {c}"), 'NumberGreater5Perc'].iat[0], 2)
        row += 1

# Table 5b
row = 226
for s in sizecats_u18_plusall:
    row += 1
    for c in range(0,18):
        ws1['B'+str(row)] = df.loc[(df['Geography'] == 'ESD') & (df['Size_Category'] == s) & (df['Characteristic'] == f"Age {c}"), 'NumCells'].iat[0]
        ws1['C'+str(row)] = round(df.loc[(df['Geography'] == 'ESD') & (df['Size_Category'] == s) & (df['Characteristic'] == f"Age {c}"), 'MeanAbsDiff'].iat[0], 2)
        ws1['D'+str(row)] = round(df.loc[(df['Geography'] == 'ESD') & (df['Size_Category'] == s) & (df['Characteristic'] == f"Age {c}"), 'RMSE'].iat[0], 2)
        ws1['E'+str(row)] = round(df.loc[(df['Geography'] == 'ESD') & (df['Size_Category'] == s) & (df['Characteristic'] == f"Age {c}"), 'MeanAbsPercDiff'].iat[0], 2)
        ws1['F'+str(row)] = round(df.loc[(df['Geography'] == 'ESD') & (df['Size_Category'] == s) & (df['Characteristic'] == f"Age {c}"), 'CV'].iat[0], 2)
        ws1['G'+str(row)] = round(df.loc[(df['Geography'] == 'ESD') & (df['Size_Category'] == s) & (df['Characteristic'] == f"Age {c}"), 'MeanPercDiff'].iat[0], 2)
        ws1['H'+str(row)] = round(df.loc[(df['Geography'] == 'ESD') & (df['Size_Category'] == s) & (df['Characteristic'] == f"Age {c}"), 'NumberGreater5Perc'].iat[0], 2)
        row += 1

# Table 5c
row = 307
for s in sizecats_u18_plusall:
    row += 1
    for c in range(0,18):
        ws1['B'+str(row)] = df.loc[(df['Geography'] == 'SSD') & (df['Size_Category'] == s) & (df['Characteristic'] == f"Age {c}"), 'NumCells'].iat[0]
        ws1['C'+str(row)] = round(df.loc[(df['Geography'] == 'SSD') & (df['Size_Category'] == s) & (df['Characteristic'] == f"Age {c}"), 'MeanAbsDiff'].iat[0], 2)
        ws1['D'+str(row)] = round(df.loc[(df['Geography'] == 'SSD') & (df['Size_Category'] == s) & (df['Characteristic'] == f"Age {c}"), 'RMSE'].iat[0], 2)
        ws1['E'+str(row)] = round(df.loc[(df['Geography'] == 'SSD') & (df['Size_Category'] == s) & (df['Characteristic'] == f"Age {c}"), 'MeanAbsPercDiff'].iat[0], 2)
        ws1['F'+str(row)] = round(df.loc[(df['Geography'] == 'SSD') & (df['Size_Category'] == s) & (df['Characteristic'] == f"Age {c}"), 'CV'].iat[0], 2)
        ws1['G'+str(row)] = round(df.loc[(df['Geography'] == 'SSD') & (df['Size_Category'] == s) & (df['Characteristic'] == f"Age {c}"), 'MeanPercDiff'].iat[0], 2)
        ws1['H'+str(row)] = round(df.loc[(df['Geography'] == 'SSD') & (df['Size_Category'] == s) & (df['Characteristic'] == f"Age {c}"), 'NumberGreater5Perc'].iat[0], 2)
        row += 1

# Table 5d
row = 388
for s in sizecats_u18_plusall:
    row += 1
    for c in range(0,18):
        ws1['B'+str(row)] = df.loc[(df['Geography'] == 'USD') & (df['Size_Category'] == s) & (df['Characteristic'] == f"Age {c}"), 'NumCells'].iat[0]
        ws1['C'+str(row)] = round(df.loc[(df['Geography'] == 'USD') & (df['Size_Category'] == s) & (df['Characteristic'] == f"Age {c}"), 'MeanAbsDiff'].iat[0], 2)
        ws1['D'+str(row)] = round(df.loc[(df['Geography'] == 'USD') & (df['Size_Category'] == s) & (df['Characteristic'] == f"Age {c}"), 'RMSE'].iat[0], 2)
        ws1['E'+str(row)] = round(df.loc[(df['Geography'] == 'USD') & (df['Size_Category'] == s) & (df['Characteristic'] == f"Age {c}"), 'MeanAbsPercDiff'].iat[0], 2)
        ws1['F'+str(row)] = round(df.loc[(df['Geography'] == 'USD') & (df['Size_Category'] == s) & (df['Characteristic'] == f"Age {c}"), 'CV'].iat[0], 2)
        ws1['G'+str(row)] = round(df.loc[(df['Geography'] == 'USD') & (df['Size_Category'] == s) & (df['Characteristic'] == f"Age {c}"), 'MeanPercDiff'].iat[0], 2)
        ws1['H'+str(row)] = round(df.loc[(df['Geography'] == 'USD') & (df['Size_Category'] == s) & (df['Characteristic'] == f"Age {c}"), 'NumberGreater5Perc'].iat[0], 2)
        row += 1

# Table 6
row = 470
ws1['B'+str(row)] = df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "TAES AIAN Alone or In Combination Nation"), 'NumCells'].iat[0]
ws1['C'+str(row)] = df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "TAES AIAN Alone or In Combination Nation"), 'TAES'].iat[0]
ws1['B'+str(row+1)] = df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "TAES AIAN Alone or In Combination Nation"), 'NumCells'].iat[0]
ws1['C'+str(row+1)] = df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "TAES AIAN Alone or In Combination Nation"), 'TAES'].iat[0]

# Table 7a
row = 478
for s in ['All','AIAN Population Size [0.0, 10.0)','AIAN Population Size [10.0, 100.0)','AIAN Population Size [100.0, inf)']:
    ws1['B'+str(row)] = df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == "AIAN Alone CountMDFltHDF"), 'NumCells'].iat[0]
    ws1['C'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == "AIAN Alone CountMDFltHDF"), 'CountMDFltHDF'].iat[0], 2)
    ws1['D'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == "AIAN Alone CountMDFltHDF"), 'MedianPctDiffWhereMDFltHDF'].iat[0], 2)
    row +=1

# Table 7b
row = 487
for s in ['All','AIAN Population Size [0.0, 10.0)','AIAN Population Size [10.0, 100.0)','AIAN Population Size [100.0, inf)']:
    ws1['B'+str(row)] = df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == "AIAN Alone CountMDFltHDF"), 'NumCells'].iat[0]
    ws1['C'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == "AIAN Alone CountMDFltHDF"), 'CountMDFltHDF'].iat[0], 2)
    ws1['D'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == "AIAN Alone CountMDFltHDF"), 'MedianPctDiffWhereMDFltHDF'].iat[0], 2)
    row +=1

# Table 7c
row = 496
for s in ['All','AIAN Population Size [0.0, 10.0)','AIAN Population Size [10.0, 100.0)','AIAN Population Size [100.0, inf)']:
    ws1['B'+str(row)] = df.loc[(df['Geography'] == 'Fed AIR') & (df['Size_Category'] == s) & (df['Characteristic'] == "AIAN Alone CountMDFltHDF"), 'NumCells'].iat[0]
    ws1['C'+str(row)] = round(df.loc[(df['Geography'] == 'Fed AIR') & (df['Size_Category'] == s) & (df['Characteristic'] == "AIAN Alone CountMDFltHDF"), 'CountMDFltHDF'].iat[0], 2)
    ws1['D'+str(row)] = round(df.loc[(df['Geography'] == 'Fed AIR') & (df['Size_Category'] == s) & (df['Characteristic'] == "AIAN Alone CountMDFltHDF"), 'MedianPctDiffWhereMDFltHDF'].iat[0], 2)
    row +=1

# Table 7d
row = 505
for s in ['All','AIAN Population Size [0.0, 10.0)','AIAN Population Size [10.0, 100.0)','AIAN Population Size [100.0, inf)']:
    ws1['B'+str(row)] = df.loc[(df['Geography'] == 'OTSA') & (df['Size_Category'] == s) & (df['Characteristic'] == "AIAN Alone CountMDFltHDF"), 'NumCells'].iat[0]
    ws1['C'+str(row)] = round(df.loc[(df['Geography'] == 'OTSA') & (df['Size_Category'] == s) & (df['Characteristic'] == "AIAN Alone CountMDFltHDF"), 'CountMDFltHDF'].iat[0], 2)
    ws1['D'+str(row)] = round(df.loc[(df['Geography'] == 'OTSA') & (df['Size_Category'] == s) & (df['Characteristic'] == "AIAN Alone CountMDFltHDF"), 'MedianPctDiffWhereMDFltHDF'].iat[0], 2)
    row +=1

# Table 7e
row = 514
for s in ['All','AIAN Population Size [0.0, 10.0)','AIAN Population Size [10.0, 100.0)','AIAN Population Size [100.0, inf)']:
    ws1['B'+str(row)] = df.loc[(df['Geography'] == 'ANVSA') & (df['Size_Category'] == s) & (df['Characteristic'] == "AIAN Alone CountMDFltHDF"), 'NumCells'].iat[0]
    ws1['C'+str(row)] = round(df.loc[(df['Geography'] == 'ANVSA') & (df['Size_Category'] == s) & (df['Characteristic'] == "AIAN Alone CountMDFltHDF"), 'CountMDFltHDF'].iat[0], 2)
    ws1['D'+str(row)] = round(df.loc[(df['Geography'] == 'ANVSA') & (df['Size_Category'] == s) & (df['Characteristic'] == "AIAN Alone CountMDFltHDF"), 'MedianPctDiffWhereMDFltHDF'].iat[0], 2)
    row +=1

# Table 8a
row = 523
for s in ['All','NHPI Population Size [0.0, 10.0)','NHPI Population Size [10.0, 100.0)','NHPI Population Size [100.0, inf)']:
    ws1['B'+str(row)] = df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == "NHPI Alone CountMDFltHDF"), 'NumCells'].iat[0]
    ws1['C'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == "NHPI Alone CountMDFltHDF"), 'CountMDFltHDF'].iat[0], 2)
    ws1['D'+str(row)] = round(df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == s) & (df['Characteristic'] == "NHPI Alone CountMDFltHDF"), 'MedianPctDiffWhereMDFltHDF'].iat[0], 2)
    row +=1

# Table 8b
row = 532
for s in ['All','NHPI Population Size [0.0, 10.0)','NHPI Population Size [10.0, 100.0)','NHPI Population Size [100.0, inf)']:
    ws1['B'+str(row)] = df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == "NHPI Alone CountMDFltHDF"), 'NumCells'].iat[0]
    ws1['C'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == "NHPI Alone CountMDFltHDF"), 'CountMDFltHDF'].iat[0], 2)
    ws1['D'+str(row)] = round(df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == s) & (df['Characteristic'] == "NHPI Alone CountMDFltHDF"), 'MedianPctDiffWhereMDFltHDF'].iat[0], 2)
    row +=1

# Table 9
row = 542
ws1['B'+str(row)] = df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "AIAN Alone Or In Combination"), 'NumCells'].iat[0]
ws1['C'+str(row)] = round(df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "AIAN Alone Or In Combination"), 'Number100PlusMDFLessThan20HDF'].iat[0], 2)
ws1['D'+str(row)] = round(df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "AIAN Alone Or In Combination"), 'NumberLessThan20MDF100PlusHDF'].iat[0], 2)

# Table 10
row = 548
ws1['B'+str(row)] = df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "NHPI Alone"), 'NumCells'].iat[0]
ws1['C'+str(row)] = round(df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "NHPI Alone"), 'Number100PlusMDFLessThan20HDF'].iat[0], 2)
ws1['D'+str(row)] = round(df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "NHPI Alone"), 'NumberLessThan20MDF100PlusHDF'].iat[0], 2)

if unitmetricsready:
    # Table 11
    row = 555
    ws1['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumCells'].iat[0]
    ws1['C'+str(row)] = dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumberNinetyPlusOnMDFNotHDF'].iat[0]
    ws1['D'+str(row)] = dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumberBtw2Pts5Pts'].iat[0]
    ws1['E'+str(row)] = dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumberGreater5Pts'].iat[0]
    ws1['B'+str(row+1)] = dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumCells'].iat[0]
    ws1['C'+str(row+1)] = dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumberNinetyPlusOnMDFNotHDF'].iat[0]
    ws1['D'+str(row+1)] = dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumberBtw2Pts5Pts'].iat[0]
    ws1['E'+str(row+1)] = dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumberGreater5Pts'].iat[0]
    ws1['B'+str(row+2)] = dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumCells'].iat[0]
    ws1['C'+str(row+2)] = dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumberNinetyPlusOnMDFNotHDF'].iat[0]
    ws1['D'+str(row+2)] = dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumberBtw2Pts5Pts'].iat[0]
    ws1['E'+str(row+2)] = dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumberGreater5Pts'].iat[0]


# Table 12
row = 564
ws1['B'+str(row)] = df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Cross 50000"), 'NumCells'].iat[0]
ws1['C'+str(row)] = df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Cross 50000"), 'NumberHDFlt50kMDFgt50k'].iat[0]
ws1['D'+str(row)] = df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Cross 50000"), 'NumberHDFgt50kMDFlt50k'].iat[0]
ws1['B'+str(row+1)] = df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Cross 50000"), 'NumCells'].iat[0]
ws1['C'+str(row+1)] = df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Cross 50000"), 'NumberHDFlt50kMDFgt50k'].iat[0]
ws1['D'+str(row+1)] = df.loc[(df['Geography'] == 'Place') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Cross 50000"), 'NumberHDFgt50kMDFlt50k'].iat[0]
ws1['B'+str(row+2)] = df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Cross 50000"), 'NumCells'].iat[0]
ws1['C'+str(row+2)] = df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Cross 50000"), 'NumberHDFlt50kMDFgt50k'].iat[0]
ws1['D'+str(row+2)] = df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Cross 50000"), 'NumberHDFgt50kMDFlt50k'].iat[0]

ws2 = wb["Full DHC Use Cases"]

if unitmetricsready:
    # Table 1a
    row = 7
    for c in ['Owned with a mortgage','Owned free and clear', 'Rented']:
        ws2['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == c), 'NumCells'].iat[0]
        ws2['C'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == c), 'MeanAbsDiff'].iat[0], 2)
        ws2['D'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == c), 'RMSE'].iat[0], 2)
        ws2['E'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == c), 'MeanAbsPercDiff'].iat[0], 2)
        ws2['F'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == c), 'CV'].iat[0], 2)
        ws2['G'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == c), 'MeanPercDiff'].iat[0], 2)
        ws2['H'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == c), 'NumberGreater5Perc'].iat[0], 2)
        row +=1
    # Table 1b
    row = 15
    for s in sizecats_plusall:
        row += 1
        for c in ['Owned with a mortgage','Owned free and clear', 'Rented']:
            ws2['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == c), 'NumCells'].iat[0]
            ws2['C'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == c), 'MeanAbsDiff'].iat[0], 2)
            ws2['D'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == c), 'RMSE'].iat[0], 2)
            ws2['E'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == c), 'MeanAbsPercDiff'].iat[0], 2)
            ws2['F'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == c), 'CV'].iat[0], 2)
            ws2['G'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == c), 'MeanPercDiff'].iat[0], 2)
            ws2['H'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == c), 'NumberGreater5Perc'].iat[0], 2)
            row += 1
    # Table 1c
    row = 48
    for s in incplacesizecats_plusall:
        row += 1
        for c in ['Owned with a mortgage','Owned free and clear', 'Rented']:
            ws2['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == c), 'NumCells'].iat[0]
            ws2['C'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == c), 'MeanAbsDiff'].iat[0], 2)
            ws2['D'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == c), 'RMSE'].iat[0], 2)
            ws2['E'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == c), 'MeanAbsPercDiff'].iat[0], 2)
            ws2['F'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == c), 'CV'].iat[0], 2)
            ws2['G'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == c), 'MeanPercDiff'].iat[0], 2)
            ws2['H'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == c), 'NumberGreater5Perc'].iat[0], 2)
            row += 1
    # Table 1d
    row = 85
    for c in ['Owned with a mortgage','Owned free and clear', 'Rented']:
        ws2['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == c), 'NumCells'].iat[0]
        ws2['C'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == c), 'MeanAbsDiff'].iat[0], 2)
        ws2['D'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == c), 'RMSE'].iat[0], 2)
        ws2['E'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == c), 'NumberGreater5Perc'].iat[0], 2)
        row += 1
    #
    # Table 2a
    row = 93
    for t in ["Owner-Occupied","Renter-Occupied"]:
        row += 1
        for c in ['Householder 15 to 24 years','Householder 25 to 34 years', 'Householder 35 to 54 years', 'Householder 55 to 64 years', 'Householder 65 years and over']:
            ws2['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{t} {c}"), 'NumCells'].iat[0]
            ws2['C'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{t} {c}"), 'MeanAbsDiff'].iat[0], 2)
            ws2['D'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{t} {c}"), 'RMSE'].iat[0], 2)
            ws2['E'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{t} {c}"), 'MeanAbsPercDiff'].iat[0], 2)
            ws2['F'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{t} {c}"), 'CV'].iat[0], 2)
            ws2['G'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{t} {c}"), 'MeanPercDiff'].iat[0], 2)
            ws2['H'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{t} {c}"), 'NumberGreater5Perc'].iat[0], 2)
            row += 1
    # Table 2b
    row = 110
    for t in ["Owner-Occupied","Renter-Occupied"]:
        row += 1
        for c in ['Householder 15 to 24 years','Householder 25 to 34 years', 'Householder 35 to 54 years', 'Householder 55 to 64 years', 'Householder 65 years and over']:
            ws2['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{t} {c}"), 'NumCells'].iat[0]
            ws2['C'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{t} {c}"), 'MeanAbsDiff'].iat[0], 2)
            ws2['D'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{t} {c}"), 'RMSE'].iat[0], 2)
            ws2['E'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{t} {c}"), 'MeanAbsPercDiff'].iat[0], 2)
            ws2['F'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{t} {c}"), 'CV'].iat[0], 2)
            ws2['G'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{t} {c}"), 'MeanPercDiff'].iat[0], 2)
            ws2['H'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{t} {c}"), 'NumberGreater5Perc'].iat[0], 2)
            row += 1
    # Table 2c
    row = 127
    for t in ["Owner-Occupied","Renter-Occupied"]:
        row += 1
        for c in ['Householder 15 to 24 years','Householder 25 to 34 years', 'Householder 35 to 54 years', 'Householder 55 to 64 years', 'Householder 65 years and over']:
            ws2['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{t} {c}"), 'NumCells'].iat[0]
            ws2['C'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{t} {c}"), 'MeanAbsDiff'].iat[0], 2)
            ws2['D'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{t} {c}"), 'RMSE'].iat[0], 2)
            ws2['E'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{t} {c}"), 'MeanAbsPercDiff'].iat[0], 2)
            ws2['F'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{t} {c}"), 'CV'].iat[0], 2)
            ws2['G'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{t} {c}"), 'MeanPercDiff'].iat[0], 2)
            ws2['H'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{t} {c}"), 'NumberGreater5Perc'].iat[0], 2)
            row += 1
    # Table 3a
    row = 144
    for t in ["Owner-Occupied","Renter-Occupied"]:
        row += 1
        for c in ['Hispanic', 'Not Hispanic']:
            ws2['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{t} {c}"), 'NumCells'].iat[0]
            ws2['C'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{t} {c}"), 'MeanAbsDiff'].iat[0], 2)
            ws2['D'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{t} {c}"), 'RMSE'].iat[0], 2)
            ws2['E'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{t} {c}"), 'MeanAbsPercDiff'].iat[0], 2)
            ws2['F'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{t} {c}"), 'CV'].iat[0], 2)
            ws2['G'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{t} {c}"), 'MeanPercDiff'].iat[0], 2)
            ws2['H'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{t} {c}"), 'NumberGreater5Perc'].iat[0], 2)
            row += 1
    # Table 3b
    row = 155
    for t in ["Owner-Occupied","Renter-Occupied"]:
        row += 1
        for c in ['Hispanic', 'Not Hispanic']:
            ws2['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{t} {c}"), 'NumCells'].iat[0]
            ws2['C'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{t} {c}"), 'MeanAbsDiff'].iat[0], 2)
            ws2['D'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{t} {c}"), 'RMSE'].iat[0], 2)
            ws2['E'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{t} {c}"), 'MeanAbsPercDiff'].iat[0], 2)
            ws2['F'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{t} {c}"), 'CV'].iat[0], 2)
            ws2['G'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{t} {c}"), 'MeanPercDiff'].iat[0], 2)
            ws2['H'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{t} {c}"), 'NumberGreater5Perc'].iat[0], 2)
            row += 1
    # Table 3c
    row = 166
    for t in ["Owner-Occupied","Renter-Occupied"]:
        row += 1
        for c in ['Hispanic', 'Not Hispanic']:
            ws2['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{t} {c}"), 'NumCells'].iat[0]
            ws2['C'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{t} {c}"), 'MeanAbsDiff'].iat[0], 2)
            ws2['D'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{t} {c}"), 'RMSE'].iat[0], 2)
            ws2['E'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{t} {c}"), 'MeanAbsPercDiff'].iat[0], 2)
            ws2['F'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{t} {c}"), 'CV'].iat[0], 2)
            ws2['G'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{t} {c}"), 'MeanPercDiff'].iat[0], 2)
            ws2['H'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{t} {c}"), 'NumberGreater5Perc'].iat[0], 2)
            row += 1
    # Table 4a
    row = 177
    for t in ["Owner-Occupied","Renter-Occupied"]:
        row += 1
        for c in list(racealonedict.values()):
            ws2['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{t} {c}"), 'NumCells'].iat[0]
            ws2['C'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{t} {c}"), 'MeanAbsDiff'].iat[0], 2)
            ws2['D'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{t} {c}"), 'RMSE'].iat[0], 2)
            ws2['E'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{t} {c}"), 'MeanAbsPercDiff'].iat[0], 2)
            ws2['F'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{t} {c}"), 'CV'].iat[0], 2)
            ws2['G'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{t} {c}"), 'MeanPercDiff'].iat[0], 2)
            ws2['H'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{t} {c}"), 'NumberGreater5Perc'].iat[0], 2)
            row += 1
    # Table 4b
    row = 198
    for t in ["Owner-Occupied","Renter-Occupied"]:
        row += 1
        for c in list(racealonedict.values()):
            ws2['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{t} {c}"), 'NumCells'].iat[0]
            ws2['C'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{t} {c}"), 'MeanAbsDiff'].iat[0], 2)
            ws2['D'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{t} {c}"), 'RMSE'].iat[0], 2)
            ws2['E'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{t} {c}"), 'MeanAbsPercDiff'].iat[0], 2)
            ws2['F'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{t} {c}"), 'CV'].iat[0], 2)
            ws2['G'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{t} {c}"), 'MeanPercDiff'].iat[0], 2)
            ws2['H'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{t} {c}"), 'NumberGreater5Perc'].iat[0], 2)
            row += 1
    # Table 4c
    row = 219
    for t in ["Owner-Occupied","Renter-Occupied"]:
        row += 1
        for c in list(racealonedict.values()):
            ws2['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{t} {c}"), 'NumCells'].iat[0]
            ws2['C'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{t} {c}"), 'MeanAbsDiff'].iat[0], 2)
            ws2['D'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{t} {c}"), 'RMSE'].iat[0], 2)
            ws2['E'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{t} {c}"), 'MeanAbsPercDiff'].iat[0], 2)
            ws2['F'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{t} {c}"), 'CV'].iat[0], 2)
            ws2['G'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{t} {c}"), 'MeanPercDiff'].iat[0], 2)
            ws2['H'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{t} {c}"), 'NumberGreater5Perc'].iat[0], 2)
            row += 1
    # Table 5a
    row = 240
    for c in list(vacantdict.values()):
        ws2['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == c), 'NumCells'].iat[0]
        ws2['C'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == c), 'MeanAbsDiff'].iat[0], 2)
        ws2['D'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == c), 'RMSE'].iat[0], 2)
        ws2['E'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == c), 'MeanAbsPercDiff'].iat[0], 2)
        ws2['F'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == c), 'CV'].iat[0], 2)
        ws2['G'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == c), 'MeanPercDiff'].iat[0], 2)
        ws2['H'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == c), 'NumberGreater5Perc'].iat[0], 2)
        row += 1
    # Table 5b
    row = 252
    for s in sizecats_plusall:
        row += 1
        for c in list(vacantdict.values()):
            ws2['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == c), 'NumCells'].iat[0]
            ws2['C'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == c), 'MeanAbsDiff'].iat[0], 2)
            ws2['D'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == c), 'RMSE'].iat[0], 2)
            ws2['E'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == c), 'MeanAbsPercDiff'].iat[0], 2)
            ws2['F'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == c), 'CV'].iat[0], 2)
            ws2['G'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == c), 'MeanPercDiff'].iat[0], 2)
            ws2['H'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == c), 'NumberGreater5Perc'].iat[0], 2)
            row += 1
    # Table 5c
    row = 313
    for s in incplacesizecats_plusall:
        row += 1
        for c in list(vacantdict.values()):
            ws2['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == c), 'NumCells'].iat[0]
            ws2['C'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == c), 'MeanAbsDiff'].iat[0], 2)
            ws2['D'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == c), 'RMSE'].iat[0], 2)
            ws2['E'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == c), 'MeanAbsPercDiff'].iat[0], 2)
            ws2['F'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == c), 'CV'].iat[0], 2)
            ws2['G'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == c), 'MeanPercDiff'].iat[0], 2)
            ws2['H'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == c), 'NumberGreater5Perc'].iat[0], 2)
            row += 1
    # Table 5d
    row = 382
    for c in list(vacantdict.values()):
        ws2['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == c), 'NumCells'].iat[0]
        ws2['C'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == c), 'MeanAbsDiff'].iat[0], 2)
        ws2['D'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == c), 'RMSE'].iat[0], 2)
        ws2['E'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == c), 'NumberGreater5Perc'].iat[0], 2)
        row += 1
    # Table 6
    row = 394
    ws2['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == "Change in Largest Vacancy Category"), 'NumCells'].iat[0]
    ws2['C'+str(row)] = dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == "Change in Largest Vacancy Category"), 'LargestVACS_DontMatch'].iat[0]
    # Table 7a
    row = 400
    for c in range(1,8):
        ws2['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"HHSize {c}"), 'NumCells'].iat[0]
        ws2['C'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"HHSize {c}"), 'MeanAbsDiff'].iat[0], 2)
        ws2['D'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"HHSize {c}"), 'RMSE'].iat[0], 2)
        ws2['E'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"HHSize {c}"), 'MeanAbsPercDiff'].iat[0], 2)
        ws2['F'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"HHSize {c}"), 'CV'].iat[0], 2)
        ws2['G'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"HHSize {c}"), 'MeanPercDiff'].iat[0], 2)
        ws2['H'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"HHSize {c}"), 'NumberGreater5Perc'].iat[0], 2)
        row += 1
    # Table 7b
    row = 412
    for s in sizecats_plusall:
        row += 1
        for c in range(1,8):
            ws2['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == f"HHSize {c}"), 'NumCells'].iat[0]
            ws2['C'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == f"HHSize {c}"), 'MeanAbsDiff'].iat[0], 2)
            ws2['D'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == f"HHSize {c}"), 'RMSE'].iat[0], 2)
            ws2['E'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == f"HHSize {c}"), 'MeanAbsPercDiff'].iat[0], 2)
            ws2['F'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == f"HHSize {c}"), 'CV'].iat[0], 2)
            ws2['G'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == f"HHSize {c}"), 'MeanPercDiff'].iat[0], 2)
            ws2['H'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == f"HHSize {c}"), 'NumberGreater5Perc'].iat[0], 2)
            row += 1
    # Table 7c
    row = 473
    for s in incplacesizecats_plusall:
        row += 1
        for c in range(1,8):
            ws2['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == f"HHSize {c}"), 'NumCells'].iat[0]
            ws2['C'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == f"HHSize {c}"), 'MeanAbsDiff'].iat[0], 2)
            ws2['D'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == f"HHSize {c}"), 'RMSE'].iat[0], 2)
            ws2['E'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == f"HHSize {c}"), 'MeanAbsPercDiff'].iat[0], 2)
            ws2['F'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == f"HHSize {c}"), 'CV'].iat[0], 2)
            ws2['G'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == f"HHSize {c}"), 'MeanPercDiff'].iat[0], 2)
            ws2['H'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == f"HHSize {c}"), 'NumberGreater5Perc'].iat[0], 2)
            row += 1
    # Table 8a
    row = 542
    ws2['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == "Aged 65+ Living Alone"), 'NumCells'].iat[0]
    ws2['C'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == "Aged 65+ Living Alone"), 'MeanAbsDiff'].iat[0], 2)
    ws2['D'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == "Aged 65+ Living Alone"), 'RMSE'].iat[0], 2)
    ws2['E'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == "Aged 65+ Living Alone"), 'MeanAbsPercDiff'].iat[0], 2)
    ws2['F'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == "Aged 65+ Living Alone"), 'CV'].iat[0], 2)
    ws2['G'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == "Aged 65+ Living Alone"), 'MeanPercDiff'].iat[0], 2)
    ws2['H'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == "Aged 65+ Living Alone"), 'NumberGreater5Perc'].iat[0], 2)
    # Table 8b
    row = 548
    for s in sizecats_plusall:
        ws2['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == "Aged 65+ Living Alone"), 'NumCells'].iat[0]
        ws2['C'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == "Aged 65+ Living Alone"), 'MeanAbsDiff'].iat[0], 2)
        ws2['D'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == "Aged 65+ Living Alone"), 'RMSE'].iat[0], 2)
        ws2['E'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == "Aged 65+ Living Alone"), 'MeanAbsPercDiff'].iat[0], 2)
        ws2['F'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == "Aged 65+ Living Alone"), 'CV'].iat[0], 2)
        ws2['G'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == "Aged 65+ Living Alone"), 'MeanPercDiff'].iat[0], 2)
        ws2['H'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == "Aged 65+ Living Alone"), 'NumberGreater5Perc'].iat[0], 2)
        row += 1
    # Table 8c
    row = 560
    for s in incplacesizecats_plusall:
        ws2['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == "Aged 65+ Living Alone"), 'NumCells'].iat[0]
        ws2['C'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == "Aged 65+ Living Alone"), 'MeanAbsDiff'].iat[0], 2)
        ws2['D'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == "Aged 65+ Living Alone"), 'RMSE'].iat[0], 2)
        ws2['E'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == "Aged 65+ Living Alone"), 'MeanAbsPercDiff'].iat[0], 2)
        ws2['F'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == "Aged 65+ Living Alone"), 'CV'].iat[0], 2)
        ws2['G'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == "Aged 65+ Living Alone"), 'MeanPercDiff'].iat[0], 2)
        ws2['H'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == "Aged 65+ Living Alone"), 'NumberGreater5Perc'].iat[0], 2)
        row += 1
    # Table 9a
    row = 573
    for c in ['Hispanic', 'Non-Hispanic']:
        ws2['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"Multigenerational {c} Householder"), 'NumCells'].iat[0]
        ws2['C'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"Multigenerational {c} Householder"), 'MeanAbsDiff'].iat[0], 2)
        ws2['D'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"Multigenerational {c} Householder"), 'RMSE'].iat[0], 2)
        ws2['E'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"Multigenerational {c} Householder"), 'MeanAbsPercDiff'].iat[0], 2)
        ws2['F'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"Multigenerational {c} Householder"), 'CV'].iat[0], 2)
        ws2['G'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"Multigenerational {c} Householder"), 'MeanPercDiff'].iat[0], 2)
        ws2['H'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"Multigenerational {c} Householder"), 'NumberGreater5Perc'].iat[0], 2)
        row += 1
    # Table 9b
    row = 580
    for s in sizecats_plusall:
        row += 1
        for c in ['Hispanic', 'Non-Hispanic']:
            ws2['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == f"Multigenerational {c} Householder"), 'NumCells'].iat[0]
            ws2['C'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == f"Multigenerational {c} Householder"), 'MeanAbsDiff'].iat[0], 2)
            ws2['D'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == f"Multigenerational {c} Householder"), 'RMSE'].iat[0], 2)
            ws2['E'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == f"Multigenerational {c} Householder"), 'MeanAbsPercDiff'].iat[0], 2)
            ws2['F'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == f"Multigenerational {c} Householder"), 'CV'].iat[0], 2)
            ws2['G'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == f"Multigenerational {c} Householder"), 'MeanPercDiff'].iat[0], 2)
            ws2['H'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == f"Multigenerational {c} Householder"), 'NumberGreater5Perc'].iat[0], 2)
            row += 1
    # Table 9c
    row = 606
    for s in incplacesizecats_plusall:
        row += 1
        for c in ['Hispanic', 'Non-Hispanic']:
            ws2['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == f"Multigenerational {c} Householder"), 'NumCells'].iat[0]
            ws2['C'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == f"Multigenerational {c} Householder"), 'MeanAbsDiff'].iat[0], 2)
            ws2['D'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == f"Multigenerational {c} Householder"), 'RMSE'].iat[0], 2)
            ws2['E'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == f"Multigenerational {c} Householder"), 'MeanAbsPercDiff'].iat[0], 2)
            ws2['F'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == f"Multigenerational {c} Householder"), 'CV'].iat[0], 2)
            ws2['G'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == f"Multigenerational {c} Householder"), 'MeanPercDiff'].iat[0], 2)
            ws2['H'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == f"Multigenerational {c} Householder"), 'NumberGreater5Perc'].iat[0], 2)
            row += 1
    # Table 10a
    row = 635
    for c in list(racealonedict.values()):
        ws2['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"Multigenerational Householder who is {c}"), 'NumCells'].iat[0]
        ws2['C'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"Multigenerational Householder who is {c}"), 'MeanAbsDiff'].iat[0], 2)
        ws2['D'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"Multigenerational Householder who is {c}"), 'RMSE'].iat[0], 2)
        ws2['E'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"Multigenerational Householder who is {c}"), 'MeanAbsPercDiff'].iat[0], 2)
        ws2['F'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"Multigenerational Householder who is {c}"), 'CV'].iat[0], 2)
        ws2['G'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"Multigenerational Householder who is {c}"), 'MeanPercDiff'].iat[0], 2)
        ws2['H'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"Multigenerational Householder who is {c}"), 'NumberGreater5Perc'].iat[0], 2)
        row += 1
    # Table 10b
    row = 647
    for s in sizecats_plusall:
        row += 1
        for c in list(racealonedict.values()):
            ws2['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == f"Multigenerational Householder who is {c}"), 'NumCells'].iat[0]
            ws2['C'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == f"Multigenerational Householder who is {c}"), 'MeanAbsDiff'].iat[0], 2)
            ws2['D'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == f"Multigenerational Householder who is {c}"), 'RMSE'].iat[0], 2)
            ws2['E'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == f"Multigenerational Householder who is {c}"), 'MeanAbsPercDiff'].iat[0], 2)
            ws2['F'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == f"Multigenerational Householder who is {c}"), 'CV'].iat[0], 2)
            ws2['G'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == f"Multigenerational Householder who is {c}"), 'MeanPercDiff'].iat[0], 2)
            ws2['H'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == f"Multigenerational Householder who is {c}"), 'NumberGreater5Perc'].iat[0], 2)
            row += 1
    # Table 10c
    row = 708
    for s in incplacesizecats_plusall:
        row += 1
        for c in list(racealonedict.values()):
            ws2['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == f"Multigenerational Householder who is {c}"), 'NumCells'].iat[0]
            ws2['C'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == f"Multigenerational Householder who is {c}"), 'MeanAbsDiff'].iat[0], 2)
            ws2['D'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == f"Multigenerational Householder who is {c}"), 'RMSE'].iat[0], 2)
            ws2['E'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == f"Multigenerational Householder who is {c}"), 'MeanAbsPercDiff'].iat[0], 2)
            ws2['F'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == f"Multigenerational Householder who is {c}"), 'CV'].iat[0], 2)
            ws2['G'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == f"Multigenerational Householder who is {c}"), 'MeanPercDiff'].iat[0], 2)
            ws2['H'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == f"Multigenerational Householder who is {c}"), 'NumberGreater5Perc'].iat[0], 2)
            row += 1
    # Table 11a
    row = 777
    for c in list(hhtypeownchildlt18dict.values()):
        ws2['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == c), 'NumCells'].iat[0]
        ws2['C'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == c), 'MeanAbsDiff'].iat[0], 2)
        ws2['D'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == c), 'RMSE'].iat[0], 2)
        ws2['E'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == c), 'MeanAbsPercDiff'].iat[0], 2)
        ws2['F'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == c), 'CV'].iat[0], 2)
        ws2['G'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == c), 'MeanPercDiff'].iat[0], 2)
        ws2['H'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == c), 'NumberGreater5Perc'].iat[0], 2)
        row += 1
    # Table 11b
    row = 786
    for s in sizecats_plusall:
        row += 1
        for c in list(hhtypeownchildlt18dict.values()):
            ws2['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == c), 'NumCells'].iat[0]
            ws2['C'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == c), 'MeanAbsDiff'].iat[0], 2)
            ws2['D'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == c), 'RMSE'].iat[0], 2)
            ws2['E'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == c), 'MeanAbsPercDiff'].iat[0], 2)
            ws2['F'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == c), 'CV'].iat[0], 2)
            ws2['G'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == c), 'MeanPercDiff'].iat[0], 2)
            ws2['H'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == c), 'NumberGreater5Perc'].iat[0], 2)
            row += 1
    # Table 11c
    row = 826
    for s in incplacesizecats_plusall:
        row += 1
        for c in list(hhtypeownchildlt18dict.values()):
            ws2['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == c), 'NumCells'].iat[0]
            ws2['C'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == c), 'MeanAbsDiff'].iat[0], 2)
            ws2['D'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == c), 'RMSE'].iat[0], 2)
            ws2['E'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == c), 'MeanAbsPercDiff'].iat[0], 2)
            ws2['F'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == c), 'CV'].iat[0], 2)
            ws2['G'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == c), 'MeanPercDiff'].iat[0], 2)
            ws2['H'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == c), 'NumberGreater5Perc'].iat[0], 2)
            row += 1
    # Table 12a
    row = 871
    ws2['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == "Presence of Own Children Under 6"), 'NumCells'].iat[0]
    ws2['C'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == "Presence of Own Children Under 6"), 'MeanAbsDiff'].iat[0], 2)
    ws2['D'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == "Presence of Own Children Under 6"), 'RMSE'].iat[0], 2)
    ws2['E'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == "Presence of Own Children Under 6"), 'MeanAbsPercDiff'].iat[0], 2)
    ws2['F'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == "Presence of Own Children Under 6"), 'CV'].iat[0], 2)
    ws2['G'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == "Presence of Own Children Under 6"), 'MeanPercDiff'].iat[0], 2)
    ws2['H'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == "Presence of Own Children Under 6"), 'NumberGreater5Perc'].iat[0], 2)
    # Table 12b
    row = 877
    for s in sizecats_plusall:
        ws2['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == "Presence of Own Children Under 6"), 'NumCells'].iat[0]
        ws2['C'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == "Presence of Own Children Under 6"), 'MeanAbsDiff'].iat[0], 2)
        ws2['D'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == "Presence of Own Children Under 6"), 'RMSE'].iat[0], 2)
        ws2['E'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == "Presence of Own Children Under 6"), 'MeanAbsPercDiff'].iat[0], 2)
        ws2['F'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == "Presence of Own Children Under 6"), 'CV'].iat[0], 2)
        ws2['G'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == "Presence of Own Children Under 6"), 'MeanPercDiff'].iat[0], 2)
        ws2['H'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == "Presence of Own Children Under 6"), 'NumberGreater5Perc'].iat[0], 2)
        row += 1
    # Table 12c
    row = 889
    for s in incplacesizecats_plusall:
        ws2['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == "Presence of Own Children Under 6"), 'NumCells'].iat[0]
        ws2['C'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == "Presence of Own Children Under 6"), 'MeanAbsDiff'].iat[0], 2)
        ws2['D'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == "Presence of Own Children Under 6"), 'RMSE'].iat[0], 2)
        ws2['E'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == "Presence of Own Children Under 6"), 'MeanAbsPercDiff'].iat[0], 2)
        ws2['F'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == "Presence of Own Children Under 6"), 'CV'].iat[0], 2)
        ws2['G'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == "Presence of Own Children Under 6"), 'MeanPercDiff'].iat[0], 2)
        ws2['H'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == s) & (dfu['Characteristic'] == "Presence of Own Children Under 6"), 'NumberGreater5Perc'].iat[0], 2)
        row += 1       
    # Table 13a
    row = 902
    for s in list(cpltdict.values()):
        ws2['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{s}"), 'NumCells'].iat[0]
        ws2['C'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{s}"), 'MeanAbsDiff'].iat[0], 2)
        ws2['D'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{s}"), 'RMSE'].iat[0], 2)
        ws2['E'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{s}"), 'MeanAbsPercDiff'].iat[0], 2)
        ws2['F'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{s}"), 'CV'].iat[0], 2)
        ws2['G'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{s}"), 'MeanPercDiff'].iat[0], 2)
        ws2['H'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{s}"), 'NumberGreater5Perc'].iat[0], 2)
        row += 1
        for c in ['Hispanic', 'Non-Hispanic']:
            ws2['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{c} {s}"), 'NumCells'].iat[0]
            ws2['C'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{c} {s}"), 'MeanAbsDiff'].iat[0], 2)
            ws2['D'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{c} {s}"), 'RMSE'].iat[0], 2)
            ws2['E'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{c} {s}"), 'MeanAbsPercDiff'].iat[0], 2)
            ws2['F'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{c} {s}"), 'CV'].iat[0], 2)
            ws2['G'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{c} {s}"), 'MeanPercDiff'].iat[0], 2)
            ws2['H'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{c} {s}"), 'NumberGreater5Perc'].iat[0], 2)
            row += 1
    # Table 13b
    row = 919
    for s in list(cpltdict.values()):
        ws2['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{s}"), 'NumCells'].iat[0]
        ws2['C'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{s}"), 'MeanAbsDiff'].iat[0], 2)
        ws2['D'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{s}"), 'RMSE'].iat[0], 2)
        ws2['E'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{s}"), 'MeanAbsPercDiff'].iat[0], 2)
        ws2['F'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{s}"), 'CV'].iat[0], 2)
        ws2['G'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{s}"), 'MeanPercDiff'].iat[0], 2)
        ws2['H'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{s}"), 'NumberGreater5Perc'].iat[0], 2)
        row += 1
        for c in ['Hispanic', 'Non-Hispanic']:
            ws2['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{c} {s}"), 'NumCells'].iat[0]
            ws2['C'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{c} {s}"), 'MeanAbsDiff'].iat[0], 2)
            ws2['D'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{c} {s}"), 'RMSE'].iat[0], 2)
            ws2['E'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{c} {s}"), 'MeanAbsPercDiff'].iat[0], 2)
            ws2['F'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{c} {s}"), 'CV'].iat[0], 2)
            ws2['G'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{c} {s}"), 'MeanPercDiff'].iat[0], 2)
            ws2['H'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{c} {s}"), 'NumberGreater5Perc'].iat[0], 2)
            row += 1
    # Table 13c
    row = 936
    for s in list(cpltdict.values()):
        ws2['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{s}"), 'NumCells'].iat[0]
        ws2['C'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{s}"), 'MeanAbsDiff'].iat[0], 2)
        ws2['D'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{s}"), 'RMSE'].iat[0], 2)
        ws2['E'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{s}"), 'MeanAbsPercDiff'].iat[0], 2)
        ws2['F'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{s}"), 'CV'].iat[0], 2)
        ws2['G'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{s}"), 'MeanPercDiff'].iat[0], 2)
        ws2['H'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{s}"), 'NumberGreater5Perc'].iat[0], 2)
        row += 1
        for c in ['Hispanic', 'Non-Hispanic']:
            ws2['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{c} {s}"), 'NumCells'].iat[0]
            ws2['C'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{c} {s}"), 'MeanAbsDiff'].iat[0], 2)
            ws2['D'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{c} {s}"), 'RMSE'].iat[0], 2)
            ws2['E'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{c} {s}"), 'MeanAbsPercDiff'].iat[0], 2)
            ws2['F'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{c} {s}"), 'CV'].iat[0], 2)
            ws2['G'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{c} {s}"), 'MeanPercDiff'].iat[0], 2)
            ws2['H'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{c} {s}"), 'NumberGreater5Perc'].iat[0], 2)
            row += 1
    # Table 14a
    row = 953
    for s in list(cpltdict.values()):
        row += 1
        for c in list(racealonedict.values()):
            ws2['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{c} {s}"), 'NumCells'].iat[0]
            ws2['C'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{c} {s}"), 'MeanAbsDiff'].iat[0], 2)
            ws2['D'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{c} {s}"), 'RMSE'].iat[0], 2)
            ws2['E'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{c} {s}"), 'MeanAbsPercDiff'].iat[0], 2)
            ws2['F'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{c} {s}"), 'CV'].iat[0], 2)
            ws2['G'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{c} {s}"), 'MeanPercDiff'].iat[0], 2)
            ws2['H'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{c} {s}"), 'NumberGreater5Perc'].iat[0], 2)
            row += 1
    # Table 14b
    row = 990
    for s in list(cpltdict.values()):
        row += 1
        for c in list(racealonedict.values()):
            ws2['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{c} {s}"), 'NumCells'].iat[0]
            ws2['C'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{c} {s}"), 'MeanAbsDiff'].iat[0], 2)
            ws2['D'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{c} {s}"), 'RMSE'].iat[0], 2)
            ws2['E'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{c} {s}"), 'MeanAbsPercDiff'].iat[0], 2)
            ws2['F'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{c} {s}"), 'CV'].iat[0], 2)
            ws2['G'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{c} {s}"), 'MeanPercDiff'].iat[0], 2)
            ws2['H'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{c} {s}"), 'NumberGreater5Perc'].iat[0], 2)
            row += 1
    # Table 14c
    row = 1027
    for s in list(cpltdict.values()):
        row += 1
        for c in list(racealonedict.values()):
            ws2['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{c} {s}"), 'NumCells'].iat[0]
            ws2['C'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{c} {s}"), 'MeanAbsDiff'].iat[0], 2)
            ws2['D'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{c} {s}"), 'RMSE'].iat[0], 2)
            ws2['E'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{c} {s}"), 'MeanAbsPercDiff'].iat[0], 2)
            ws2['F'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{c} {s}"), 'CV'].iat[0], 2)
            ws2['G'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{c} {s}"), 'MeanPercDiff'].iat[0], 2)
            ws2['H'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{c} {s}"), 'NumberGreater5Perc'].iat[0], 2)
            row += 1
    # Table 15a
    row = 1064
    for c in list(reldict.values()):
            ws2['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{c}"), 'NumCells'].iat[0]
            ws2['C'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{c}"), 'MeanAbsDiff'].iat[0], 2)
            ws2['D'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{c}"), 'RMSE'].iat[0], 2)
            ws2['E'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{c}"), 'MeanAbsPercDiff'].iat[0], 2)
            ws2['F'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{c}"), 'CV'].iat[0], 2)
            ws2['G'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{c}"), 'MeanPercDiff'].iat[0], 2)
            ws2['H'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'State') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{c}"), 'NumberGreater5Perc'].iat[0], 2)
            row += 1
    # Table 15b
    row = 1086
    for c in list(reldict.values()):
            ws2['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{c}"), 'NumCells'].iat[0]
            ws2['C'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{c}"), 'MeanAbsDiff'].iat[0], 2)
            ws2['D'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{c}"), 'RMSE'].iat[0], 2)
            ws2['E'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{c}"), 'MeanAbsPercDiff'].iat[0], 2)
            ws2['F'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{c}"), 'CV'].iat[0], 2)
            ws2['G'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{c}"), 'MeanPercDiff'].iat[0], 2)
            ws2['H'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'County') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{c}"), 'NumberGreater5Perc'].iat[0], 2)
            row += 1
    # Table 15c
    row = 1108
    for c in list(reldict.values()):
            ws2['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{c}"), 'NumCells'].iat[0]
            ws2['C'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{c}"), 'MeanAbsDiff'].iat[0], 2)
            ws2['D'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{c}"), 'RMSE'].iat[0], 2)
            ws2['E'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{c}"), 'MeanAbsPercDiff'].iat[0], 2)
            ws2['F'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{c}"), 'CV'].iat[0], 2)
            ws2['G'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{c}"), 'MeanPercDiff'].iat[0], 2)
            ws2['H'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Place') & (dfu['Size_Category'] == "All") & (dfu['Characteristic'] == f"{c}"), 'NumberGreater5Perc'].iat[0], 2)
            row += 1


if userelfile:
    # Table 15a
    row = 1064
    for c in list(reldict.values()):
            ws2['B'+str(row)] = dfrel.loc[(dfrel['Geography'] == 'State') & (dfrel['Size_Category'] == "All") & (dfrel['Characteristic'] == f"{c}"), 'NumCells'].iat[0]
            ws2['C'+str(row)] = round(dfrel.loc[(dfrel['Geography'] == 'State') & (dfrel['Size_Category'] == "All") & (dfrel['Characteristic'] == f"{c}"), 'MeanAbsDiff'].iat[0], 2)
            ws2['D'+str(row)] = round(dfrel.loc[(dfrel['Geography'] == 'State') & (dfrel['Size_Category'] == "All") & (dfrel['Characteristic'] == f"{c}"), 'RMSE'].iat[0], 2)
            ws2['E'+str(row)] = round(dfrel.loc[(dfrel['Geography'] == 'State') & (dfrel['Size_Category'] == "All") & (dfrel['Characteristic'] == f"{c}"), 'MeanAbsPercDiff'].iat[0], 2)
            ws2['F'+str(row)] = round(dfrel.loc[(dfrel['Geography'] == 'State') & (dfrel['Size_Category'] == "All") & (dfrel['Characteristic'] == f"{c}"), 'CV'].iat[0], 2)
            ws2['G'+str(row)] = round(dfrel.loc[(dfrel['Geography'] == 'State') & (dfrel['Size_Category'] == "All") & (dfrel['Characteristic'] == f"{c}"), 'MeanPercDiff'].iat[0], 2)
            ws2['H'+str(row)] = round(dfrel.loc[(dfrel['Geography'] == 'State') & (dfrel['Size_Category'] == "All") & (dfrel['Characteristic'] == f"{c}"), 'NumberGreater5Perc'].iat[0], 2)
            row += 1
    # Table 15b
    row = 1086
    for c in list(reldict.values()):
            ws2['B'+str(row)] = dfrel.loc[(dfrel['Geography'] == 'County') & (dfrel['Size_Category'] == "All") & (dfrel['Characteristic'] == f"{c}"), 'NumCells'].iat[0]
            ws2['C'+str(row)] = round(dfrel.loc[(dfrel['Geography'] == 'County') & (dfrel['Size_Category'] == "All") & (dfrel['Characteristic'] == f"{c}"), 'MeanAbsDiff'].iat[0], 2)
            ws2['D'+str(row)] = round(dfrel.loc[(dfrel['Geography'] == 'County') & (dfrel['Size_Category'] == "All") & (dfrel['Characteristic'] == f"{c}"), 'RMSE'].iat[0], 2)
            ws2['E'+str(row)] = round(dfrel.loc[(dfrel['Geography'] == 'County') & (dfrel['Size_Category'] == "All") & (dfrel['Characteristic'] == f"{c}"), 'MeanAbsPercDiff'].iat[0], 2)
            ws2['F'+str(row)] = round(dfrel.loc[(dfrel['Geography'] == 'County') & (dfrel['Size_Category'] == "All") & (dfrel['Characteristic'] == f"{c}"), 'CV'].iat[0], 2)
            ws2['G'+str(row)] = round(dfrel.loc[(dfrel['Geography'] == 'County') & (dfrel['Size_Category'] == "All") & (dfrel['Characteristic'] == f"{c}"), 'MeanPercDiff'].iat[0], 2)
            ws2['H'+str(row)] = round(dfrel.loc[(dfrel['Geography'] == 'County') & (dfrel['Size_Category'] == "All") & (dfrel['Characteristic'] == f"{c}"), 'NumberGreater5Perc'].iat[0], 2)
            row += 1
    # Table 15c
    row = 1108
    for c in list(reldict.values()):
            ws2['B'+str(row)] = dfrel.loc[(dfrel['Geography'] == 'Place') & (dfrel['Size_Category'] == "All") & (dfrel['Characteristic'] == f"{c}"), 'NumCells'].iat[0]
            ws2['C'+str(row)] = round(dfrel.loc[(dfrel['Geography'] == 'Place') & (dfrel['Size_Category'] == "All") & (dfrel['Characteristic'] == f"{c}"), 'MeanAbsDiff'].iat[0], 2)
            ws2['D'+str(row)] = round(dfrel.loc[(dfrel['Geography'] == 'Place') & (dfrel['Size_Category'] == "All") & (dfrel['Characteristic'] == f"{c}"), 'RMSE'].iat[0], 2)
            ws2['E'+str(row)] = round(dfrel.loc[(dfrel['Geography'] == 'Place') & (dfrel['Size_Category'] == "All") & (dfrel['Characteristic'] == f"{c}"), 'MeanAbsPercDiff'].iat[0], 2)
            ws2['F'+str(row)] = round(dfrel.loc[(dfrel['Geography'] == 'Place') & (dfrel['Size_Category'] == "All") & (dfrel['Characteristic'] == f"{c}"), 'CV'].iat[0], 2)
            ws2['G'+str(row)] = round(dfrel.loc[(dfrel['Geography'] == 'Place') & (dfrel['Size_Category'] == "All") & (dfrel['Characteristic'] == f"{c}"), 'MeanPercDiff'].iat[0], 2)
            ws2['H'+str(row)] = round(dfrel.loc[(dfrel['Geography'] == 'Place') & (dfrel['Size_Category'] == "All") & (dfrel['Characteristic'] == f"{c}"), 'NumberGreater5Perc'].iat[0], 2)
            row += 1

if unitmetricsready:
    # Table 15a
    row = 1064
    for c in list(reldict.values()):
            ws2['B'+str(row)] = dfrel.loc[(dfrel['Geography'] == 'State') & (dfrel['Size_Category'] == "All") & (dfrel['Characteristic'] == f"{c}"), 'NumCells'].iat[0]
            ws2['C'+str(row)] = round(dfrel.loc[(dfrel['Geography'] == 'State') & (dfrel['Size_Category'] == "All") & (dfrel['Characteristic'] == f"{c}"), 'MeanAbsDiff'].iat[0], 2)
            ws2['D'+str(row)] = round(dfrel.loc[(dfrel['Geography'] == 'State') & (dfrel['Size_Category'] == "All") & (dfrel['Characteristic'] == f"{c}"), 'RMSE'].iat[0], 2)
            ws2['E'+str(row)] = round(dfrel.loc[(dfrel['Geography'] == 'State') & (dfrel['Size_Category'] == "All") & (dfrel['Characteristic'] == f"{c}"), 'MeanAbsPercDiff'].iat[0], 2)
            ws2['F'+str(row)] = round(dfrel.loc[(dfrel['Geography'] == 'State') & (dfrel['Size_Category'] == "All") & (dfrel['Characteristic'] == f"{c}"), 'CV'].iat[0], 2)
            ws2['G'+str(row)] = round(dfrel.loc[(dfrel['Geography'] == 'State') & (dfrel['Size_Category'] == "All") & (dfrel['Characteristic'] == f"{c}"), 'MeanPercDiff'].iat[0], 2)
            ws2['H'+str(row)] = round(dfrel.loc[(dfrel['Geography'] == 'State') & (dfrel['Size_Category'] == "All") & (dfrel['Characteristic'] == f"{c}"), 'NumberGreater5Perc'].iat[0], 2)
            row += 1
    # Table 15b
    row = 1086
    for c in list(reldict.values()):
            ws2['B'+str(row)] = dfrel.loc[(dfrel['Geography'] == 'County') & (dfrel['Size_Category'] == "All") & (dfrel['Characteristic'] == f"{c}"), 'NumCells'].iat[0]
            ws2['C'+str(row)] = round(dfrel.loc[(dfrel['Geography'] == 'County') & (dfrel['Size_Category'] == "All") & (dfrel['Characteristic'] == f"{c}"), 'MeanAbsDiff'].iat[0], 2)
            ws2['D'+str(row)] = round(dfrel.loc[(dfrel['Geography'] == 'County') & (dfrel['Size_Category'] == "All") & (dfrel['Characteristic'] == f"{c}"), 'RMSE'].iat[0], 2)
            ws2['E'+str(row)] = round(dfrel.loc[(dfrel['Geography'] == 'County') & (dfrel['Size_Category'] == "All") & (dfrel['Characteristic'] == f"{c}"), 'MeanAbsPercDiff'].iat[0], 2)
            ws2['F'+str(row)] = round(dfrel.loc[(dfrel['Geography'] == 'County') & (dfrel['Size_Category'] == "All") & (dfrel['Characteristic'] == f"{c}"), 'CV'].iat[0], 2)
            ws2['G'+str(row)] = round(dfrel.loc[(dfrel['Geography'] == 'County') & (dfrel['Size_Category'] == "All") & (dfrel['Characteristic'] == f"{c}"), 'MeanPercDiff'].iat[0], 2)
            ws2['H'+str(row)] = round(dfrel.loc[(dfrel['Geography'] == 'County') & (dfrel['Size_Category'] == "All") & (dfrel['Characteristic'] == f"{c}"), 'NumberGreater5Perc'].iat[0], 2)
            row += 1
    # Table 15c
    row = 1108
    for c in list(reldict.values()):
            ws2['B'+str(row)] = dfrel.loc[(dfrel['Geography'] == 'Place') & (dfrel['Size_Category'] == "All") & (dfrel['Characteristic'] == f"{c}"), 'NumCells'].iat[0]
            ws2['C'+str(row)] = round(dfrel.loc[(dfrel['Geography'] == 'Place') & (dfrel['Size_Category'] == "All") & (dfrel['Characteristic'] == f"{c}"), 'MeanAbsDiff'].iat[0], 2)
            ws2['D'+str(row)] = round(dfrel.loc[(dfrel['Geography'] == 'Place') & (dfrel['Size_Category'] == "All") & (dfrel['Characteristic'] == f"{c}"), 'RMSE'].iat[0], 2)
            ws2['E'+str(row)] = round(dfrel.loc[(dfrel['Geography'] == 'Place') & (dfrel['Size_Category'] == "All") & (dfrel['Characteristic'] == f"{c}"), 'MeanAbsPercDiff'].iat[0], 2)
            ws2['F'+str(row)] = round(dfrel.loc[(dfrel['Geography'] == 'Place') & (dfrel['Size_Category'] == "All") & (dfrel['Characteristic'] == f"{c}"), 'CV'].iat[0], 2)
            ws2['G'+str(row)] = round(dfrel.loc[(dfrel['Geography'] == 'Place') & (dfrel['Size_Category'] == "All") & (dfrel['Characteristic'] == f"{c}"), 'MeanPercDiff'].iat[0], 2)
            ws2['H'+str(row)] = round(dfrel.loc[(dfrel['Geography'] == 'Place') & (dfrel['Size_Category'] == "All") & (dfrel['Characteristic'] == f"{c}"), 'NumberGreater5Perc'].iat[0], 2)
            row += 1

ws3 = wb["Impossible & Improbable Metrics"]
# Table 1
row = 7
ws3['B'+str(row)] = df.loc[(df['Geography'] == 'State') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population"), 'NumCells'].iat[0]
ws3['C'+str(row)] = df.loc[(df['Geography'] == 'State') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "Total Population"), 'NumberHDFneMDF'].iat[0]

if unitmetricsready:
    # Table 2
    row = 13
    ws3['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occ Households > Population"), 'NumCells'].iat[0]
    ws3['C'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occ Households > Population"), 'MDFOccHH_OutnumberPop'].iat[0])
    ws3['D'+str(row)] = round(dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "HH Size Pop > Population"), 'MDFHHSizePop_OutnumberPop'].iat[0])
    # Table 3
    row = 20
    ws3['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'Block') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "HDF Occ Units Vs Population"), 'BlocksWith1pOccHU'].iat[0]
    ws3['C'+str(row)] = dfu.loc[(dfu['Geography'] == 'Block') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "HDF Occ Units Vs Population"), 'HDFOccHH_OutnumberPop'].iat[0]
    ws3['D'+str(row)] = round(100* (dfu.loc[(dfu['Geography'] == 'Block') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "HDF Occ Units Vs Population"), 'HDFOccHH_OutnumberPop'].iat[0])/dfu.loc[(dfu['Geography'] == 'Block') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "HDF Occ Units Vs Population"), 'BlocksWith1pOccHU'].iat[0], 2)
    ws3['E'+str(row)] = dfu.loc[(dfu['Geography'] == 'Block') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "MDF Occ Units Vs Population"), 'BlocksWith1pOccHU'].iat[0]
    ws3['F'+str(row)] = dfu.loc[(dfu['Geography'] == 'Block') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "MDF Occ Units Vs Population"), 'MDFOccHH_OutnumberPop'].iat[0]
    ws3['G'+str(row)] = round(100* (dfu.loc[(dfu['Geography'] == 'Block') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "MDF Occ Units Vs Population"), 'MDFOccHH_OutnumberPop'].iat[0])/dfu.loc[(dfu['Geography'] == 'Block') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "MDF Occ Units Vs Population"), 'BlocksWith1pOccHU'].iat[0], 2)
    # Table 4
    row = 26
    ws3['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'Block') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "HDF Occ Units Vs Population"), 'BlocksWith1pPop'].iat[0]
    ws3['C'+str(row)] = dfu.loc[(dfu['Geography'] == 'Block') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "HDF Occ Units Vs Population"), 'HDFZeroOccHH_1pPop'].iat[0]
    ws3['D'+str(row)] = round(100* (dfu.loc[(dfu['Geography'] == 'Block') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "HDF Occ Units Vs Population"), 'HDFZeroOccHH_1pPop'].iat[0])/dfu.loc[(dfu['Geography'] == 'Block') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "HDF Occ Units Vs Population"), 'BlocksWith1pPop'].iat[0], 2)
    ws3['E'+str(row)] = dfu.loc[(dfu['Geography'] == 'Block') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "MDF Occ Units Vs Population"), 'BlocksWith1pPop'].iat[0]
    ws3['F'+str(row)] = dfu.loc[(dfu['Geography'] == 'Block') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "MDF Occ Units Vs Population"), 'MDFZeroOccHH_1pPop'].iat[0]
    ws3['G'+str(row)] = round(100* (dfu.loc[(dfu['Geography'] == 'Block') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "MDF Occ Units Vs Population"), 'MDFZeroOccHH_1pPop'].iat[0])/dfu.loc[(dfu['Geography'] == 'Block') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "MDF Occ Units Vs Population"), 'BlocksWith1pPop'].iat[0], 2)
    # Table 5
    row = 32
    ws3['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'Block') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "HDF Occ Units Vs Population"), 'BlocksWith1pOccHU'].iat[0]
    ws3['C'+str(row)] = dfu.loc[(dfu['Geography'] == 'Block') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "HDF Occ Units Vs Population"), 'HDFZeroPop_1pOccHH'].iat[0]
    ws3['D'+str(row)] = round(100* (dfu.loc[(dfu['Geography'] == 'Block') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "HDF Occ Units Vs Population"), 'HDFZeroPop_1pOccHH'].iat[0])/dfu.loc[(dfu['Geography'] == 'Block') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "HDF Occ Units Vs Population"), 'BlocksWith1pOccHU'].iat[0], 2)
    ws3['E'+str(row)] = dfu.loc[(dfu['Geography'] == 'Block') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "MDF Occ Units Vs Population"), 'BlocksWith1pOccHU'].iat[0]
    ws3['F'+str(row)] = dfu.loc[(dfu['Geography'] == 'Block') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "MDF Occ Units Vs Population"), 'MDFZeroPop_1pOccHH'].iat[0]
    ws3['G'+str(row)] = round(100* (dfu.loc[(dfu['Geography'] == 'Block') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "MDF Occ Units Vs Population"), 'MDFZeroPop_1pOccHH'].iat[0])/dfu.loc[(dfu['Geography'] == 'Block') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "MDF Occ Units Vs Population"), 'BlocksWith1pOccHU'].iat[0], 2)
    # Table 6
    row = 38
    ws3['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "HDF HHers>Pop By Race"), 'NumCells'].iat[0]
    ws3['C'+str(row)] = dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "HDF HHers>Pop By Race"), 'Inconsistent'].iat[0]
    ws3['D'+str(row)] = round(100* (dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "HDF HHers>Pop By Race"), 'Inconsistent'].iat[0])/dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "HDF HHers>Pop By Race"), 'NumCells'].iat[0], 2)
    ws3['E'+str(row)] = dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "MDF HHers>Pop By Race"), 'NumCells'].iat[0]
    ws3['F'+str(row)] = dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "MDF HHers>Pop By Race"), 'Inconsistent'].iat[0]
    ws3['G'+str(row)] = round(100* (dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "MDF HHers>Pop By Race"), 'Inconsistent'].iat[0])/dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "MDF HHers>Pop By Race"), 'NumCells'].iat[0], 2)
    # Table 7
    row = 44
    ws3['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "HDF HHers>Pop Hispanic"), 'NumCells'].iat[0]
    ws3['C'+str(row)] = dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "HDF HHers>Pop Hispanic"), 'Inconsistent'].iat[0]
    ws3['D'+str(row)] = round(100* (dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "HDF HHers>Pop Hispanic"), 'Inconsistent'].iat[0])/dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "HDF HHers>Pop Hispanic"), 'NumCells'].iat[0], 2)
    ws3['E'+str(row)] = dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "MDF HHers>Pop Hispanic"), 'NumCells'].iat[0]
    ws3['F'+str(row)] = dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "MDF HHers>Pop Hispanic"), 'Inconsistent'].iat[0]
    ws3['G'+str(row)] = round(100* (dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "MDF HHers>Pop Hispanic"), 'Inconsistent'].iat[0])/dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "MDF HHers>Pop Hispanic"), 'NumCells'].iat[0], 2)
    # Table 8
    row = 50
    ws3['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "HDF HHers>Pop By Age"), 'NumCells'].iat[0]
    ws3['C'+str(row)] = dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "HDF HHers>Pop By Age"), 'Inconsistent'].iat[0]
    ws3['D'+str(row)] = round(100* (dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "HDF HHers>Pop By Age"), 'Inconsistent'].iat[0])/dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "HDF HHers>Pop By Age"), 'NumCells'].iat[0], 2)
    ws3['E'+str(row)] = dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "MDF HHers>Pop By Age"), 'NumCells'].iat[0]
    ws3['F'+str(row)] = dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "MDF HHers>Pop By Age"), 'Inconsistent'].iat[0]
    ws3['G'+str(row)] = round(100* (dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "MDF HHers>Pop By Age"), 'Inconsistent'].iat[0])/dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "MDF HHers>Pop By Age"), 'NumCells'].iat[0], 2)
    # Table 9
    row = 56
    ws3['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "HDF HHers>Pop Female"), 'NumCells'].iat[0]
    ws3['C'+str(row)] = dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "HDF HHers>Pop Female"), 'Inconsistent'].iat[0]
    ws3['D'+str(row)] = round(100* (dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "HDF HHers>Pop Female"), 'Inconsistent'].iat[0])/dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "HDF HHers>Pop Female"), 'NumCells'].iat[0], 2)
    ws3['E'+str(row)] = dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "MDF HHers>Pop Female"), 'NumCells'].iat[0]
    ws3['F'+str(row)] = dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "MDF HHers>Pop Female"), 'Inconsistent'].iat[0]
    ws3['G'+str(row)] = round(100* (dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "MDF HHers>Pop Female"), 'Inconsistent'].iat[0])/dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "MDF HHers>Pop Female"), 'NumCells'].iat[0], 2)
    # Table 10
    row = 62
    ws3['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "HDF HHs>Pop Under 18"), 'NumCells'].iat[0]
    ws3['C'+str(row)] = dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "HDF HHs>Pop Under 18"), 'Inconsistent'].iat[0]
    ws3['D'+str(row)] = round(100* (dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "HDF HHs>Pop Under 18"), 'Inconsistent'].iat[0])/dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "HDF HHs>Pop Under 18"), 'NumCells'].iat[0], 2)
    ws3['E'+str(row)] = dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "MDF HHs>Pop Under 18"), 'NumCells'].iat[0]
    ws3['F'+str(row)] = dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "MDF HHs>Pop Under 18"), 'Inconsistent'].iat[0]
    ws3['G'+str(row)] = round(100* (dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "MDF HHs>Pop Under 18"), 'Inconsistent'].iat[0])/dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "MDF HHs>Pop Under 18"), 'NumCells'].iat[0], 2)
    # Table 11
    row = 68
    ws3['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "HDF OS Married HHs>Pop Male"), 'NumCells'].iat[0]
    ws3['C'+str(row)] = dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "HDF OS Married HHs>Pop Male"), 'Inconsistent'].iat[0]
    ws3['D'+str(row)] = round(100* (dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "HDF OS Married HHs>Pop Male"), 'Inconsistent'].iat[0])/dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "HDF OS Married HHs>Pop Male"), 'NumCells'].iat[0], 2)
    ws3['E'+str(row)] = dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "MDF OS Married HHs>Pop Male"), 'NumCells'].iat[0]
    ws3['F'+str(row)] = dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "MDF OS Married HHs>Pop Male"), 'Inconsistent'].iat[0]
    ws3['G'+str(row)] = round(100* (dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "MDF OS Married HHs>Pop Male"), 'Inconsistent'].iat[0])/dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "MDF OS Married HHs>Pop Male"), 'NumCells'].iat[0], 2)
    # Table 12
    row = 74
    ws3['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "HDF OS Married HHs>Pop Female"), 'NumCells'].iat[0]
    ws3['C'+str(row)] = dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "HDF OS Married HHs>Pop Female"), 'Inconsistent'].iat[0]
    ws3['D'+str(row)] = round(100* (dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "HDF OS Married HHs>Pop Female"), 'Inconsistent'].iat[0])/dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "HDF OS Married HHs>Pop Female"), 'NumCells'].iat[0], 2)
    ws3['E'+str(row)] = dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "MDF OS Married HHs>Pop Female"), 'NumCells'].iat[0]
    ws3['F'+str(row)] = dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "MDF OS Married HHs>Pop Female"), 'Inconsistent'].iat[0]
    ws3['G'+str(row)] = round(100* (dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "MDF OS Married HHs>Pop Female"), 'Inconsistent'].iat[0])/dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "MDF OS Married HHs>Pop Female"), 'NumCells'].iat[0], 2)
    # Table 13
    row = 80
    ws3['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "HDF OS Married HHs>Pop 15plus * 2"), 'NumCells'].iat[0]
    ws3['C'+str(row)] = dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "HDF OS Married HHs>Pop 15plus * 2"), 'Inconsistent'].iat[0]
    ws3['D'+str(row)] = round(100* (dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "HDF OS Married HHs>Pop 15plus * 2"), 'Inconsistent'].iat[0])/dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "HDF OS Married HHs>Pop 15plus * 2"), 'NumCells'].iat[0], 2)
    ws3['E'+str(row)] = dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "MDF OS Married HHs>Pop 15plus * 2"), 'NumCells'].iat[0]
    ws3['F'+str(row)] = dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "MDF OS Married HHs>Pop 15plus * 2"), 'Inconsistent'].iat[0]
    ws3['G'+str(row)] = round(100* (dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "MDF OS Married HHs>Pop 15plus * 2"), 'Inconsistent'].iat[0])/dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "MDF OS Married HHs>Pop 15plus * 2"), 'NumCells'].iat[0], 2)

# Table 14
row = 86
ws3['B'+str(row)] = df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "HDF Children No Moms"), 'NumCells'].iat[0]
ws3['C'+str(row)] = df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "HDF Children No Moms"), 'Inconsistent'].iat[0]
ws3['D'+str(row)] = round(100* (df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "HDF Children No Moms"), 'Inconsistent'].iat[0])/df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "HDF Children No Moms"), 'NumCells'].iat[0], 2)
ws3['E'+str(row)] = df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "MDF Children No Moms"), 'NumCells'].iat[0]
ws3['F'+str(row)] = df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "MDF Children No Moms"), 'Inconsistent'].iat[0]
ws3['G'+str(row)] = round(100* (df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "MDF Children No Moms"), 'Inconsistent'].iat[0])/df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "MDF Children No Moms"), 'NumCells'].iat[0], 2)
row += 1
ws3['B'+str(row)] = df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "HDF Children No Moms"), 'NumCells'].iat[0]
ws3['C'+str(row)] = df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "HDF Children No Moms"), 'Inconsistent'].iat[0]
ws3['D'+str(row)] = round(100* (df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "HDF Children No Moms"), 'Inconsistent'].iat[0])/df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "HDF Children No Moms"), 'NumCells'].iat[0], 2)
ws3['E'+str(row)] = df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "MDF Children No Moms"), 'NumCells'].iat[0]
ws3['F'+str(row)] = df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "MDF Children No Moms"), 'Inconsistent'].iat[0]
ws3['G'+str(row)] = round(100* (df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "MDF Children No Moms"), 'Inconsistent'].iat[0])/df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "MDF Children No Moms"), 'NumCells'].iat[0], 2)

# Table 15
row = 93
row += 2
for c in ['White Alone', 'Black Alone', 'AIAN Alone', 'Asian Alone', 'NHOPI Alone', 'Some Other Race Alone', 'Two Or More Races', 'Hispanic', 'Non-Hispanic White']:
    ws3['B'+str(row)] = df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == 'All') & (df['Characteristic'] == f"HDF Children No Moms {c}"), 'NumCells'].iat[0]
    ws3['C'+str(row)] = df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == 'All') & (df['Characteristic'] == f"HDF Children No Moms {c}"), 'Inconsistent'].iat[0]
    ws3['D'+str(row)] = round(100* (df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == 'All') & (df['Characteristic'] == f"HDF Children No Moms {c}"), 'Inconsistent'].iat[0])/df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == 'All') & (df['Characteristic'] == f"HDF Children No Moms {c}"), 'NumCells'].iat[0], 2)
    ws3['E'+str(row)] = df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == 'All') & (df['Characteristic'] == f"MDF Children No Moms {c}"), 'NumCells'].iat[0]
    ws3['F'+str(row)] = df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == 'All') & (df['Characteristic'] == f"MDF Children No Moms {c}"), 'Inconsistent'].iat[0]
    ws3['G'+str(row)] = round(100* (df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == 'All') & (df['Characteristic'] == f"MDF Children No Moms {c}"), 'Inconsistent'].iat[0])/df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == 'All') & (df['Characteristic'] == f"MDF Children No Moms {c}"), 'NumCells'].iat[0], 2)
    row += 1
row = 104
row += 2
for c in ['White Alone', 'Black Alone', 'AIAN Alone', 'Asian Alone', 'NHOPI Alone', 'Some Other Race Alone', 'Two Or More Races', 'Hispanic', 'Non-Hispanic White']:
    ws3['B'+str(row)] = df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == f"HDF Children No Moms {c}"), 'NumCells'].iat[0]
    ws3['C'+str(row)] = df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == f"HDF Children No Moms {c}"), 'Inconsistent'].iat[0]
    ws3['D'+str(row)] = round(100* (df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == f"HDF Children No Moms {c}"), 'Inconsistent'].iat[0])/df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == f"HDF Children No Moms {c}"), 'NumCells'].iat[0], 2)
    ws3['E'+str(row)] = df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == f"MDF Children No Moms {c}"), 'NumCells'].iat[0]
    ws3['F'+str(row)] = df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == f"MDF Children No Moms {c}"), 'Inconsistent'].iat[0]
    ws3['G'+str(row)] = round(100* (df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == f"MDF Children No Moms {c}"), 'Inconsistent'].iat[0])/df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == f"MDF Children No Moms {c}"), 'NumCells'].iat[0], 2)
    row += 1

# Table 16
row = 120
ws3['B'+str(row)] = df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "HDF All Same Sex"), 'NumCells'].iat[0]
ws3['C'+str(row)] = df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "HDF All Same Sex"), 'Inconsistent'].iat[0]
ws3['D'+str(row)] = round(100* (df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "HDF All Same Sex"), 'Inconsistent'].iat[0])/df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "HDF All Same Sex"), 'NumCells'].iat[0], 2)
ws3['E'+str(row)] = df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "MDF All Same Sex"), 'NumCells'].iat[0]
ws3['F'+str(row)] = df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "MDF All Same Sex"), 'Inconsistent'].iat[0]
ws3['G'+str(row)] = round(100* (df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "MDF All Same Sex"), 'Inconsistent'].iat[0])/df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "MDF All Same Sex"), 'NumCells'].iat[0], 2)

# Table 17
row = 126
ws3['B'+str(row)] = df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "HDF Zero Age"), 'NumCells'].iat[0]
ws3['C'+str(row)] = df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "HDF Zero Age"), 'Inconsistent'].iat[0]
ws3['D'+str(row)] = round(100* (df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "HDF Zero Age"), 'Inconsistent'].iat[0])/df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "HDF Zero Age"), 'NumCells'].iat[0], 2)
ws3['E'+str(row)] = df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "MDF Zero Age"), 'NumCells'].iat[0]
ws3['F'+str(row)] = df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "MDF Zero Age"), 'Inconsistent'].iat[0]
ws3['G'+str(row)] = round(100* (df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "MDF Zero Age"), 'Inconsistent'].iat[0])/df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "MDF Zero Age"), 'NumCells'].iat[0], 2)

# Table 18
row = 132
ws3['B'+str(row)] = df.loc[(df['Geography'] == 'Block') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "HDF Everyone Under 18"), 'NumCells'].iat[0]
ws3['C'+str(row)] = df.loc[(df['Geography'] == 'Block') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "HDF Everyone Under 18"), 'Inconsistent'].iat[0]
ws3['D'+str(row)] = round(100* (df.loc[(df['Geography'] == 'Block') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "HDF Everyone Under 18"), 'Inconsistent'].iat[0])/df.loc[(df['Geography'] == 'Block') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "HDF Everyone Under 18"), 'NumCells'].iat[0], 2)
ws3['E'+str(row)] = df.loc[(df['Geography'] == 'Block') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "MDF Everyone Under 18"), 'NumCells'].iat[0]
ws3['F'+str(row)] = df.loc[(df['Geography'] == 'Block') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "MDF Everyone Under 18"), 'Inconsistent'].iat[0]
ws3['G'+str(row)] = round(100* (df.loc[(df['Geography'] == 'Block') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "MDF Everyone Under 18"), 'Inconsistent'].iat[0])/df.loc[(df['Geography'] == 'Block') & (df['Size_Category'] == 'All') & (df['Characteristic'] == "MDF Everyone Under 18"), 'NumCells'].iat[0], 2)

if unitmetricsready:
    # Table 19
    row = 138
    ws3['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'Block') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "HDF Occ Units Vs Population"), 'BlocksWith1pOccHU'].iat[0]
    ws3['C'+str(row)] = dfu.loc[(dfu['Geography'] == 'Block') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "HDF Occ Units Vs Population"), 'HDFgt10PPH'].iat[0]
    ws3['D'+str(row)] = round(100* (dfu.loc[(dfu['Geography'] == 'Block') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "HDF Occ Units Vs Population"), 'HDFgt10PPH'].iat[0])/dfu.loc[(dfu['Geography'] == 'Block') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "HDF Occ Units Vs Population"), 'BlocksWith1pOccHU'].iat[0], 2)
    ws3['E'+str(row)] = dfu.loc[(dfu['Geography'] == 'Block') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "MDF Occ Units Vs Population"), 'BlocksWith1pOccHU'].iat[0]
    ws3['F'+str(row)] = dfu.loc[(dfu['Geography'] == 'Block') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "MDF Occ Units Vs Population"), 'MDFgt10PPH'].iat[0]
    ws3['G'+str(row)] = round(100* (dfu.loc[(dfu['Geography'] == 'Block') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "MDF Occ Units Vs Population"), 'MDFgt10PPH'].iat[0])/dfu.loc[(dfu['Geography'] == 'Block') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "MDF Occ Units Vs Population"), 'BlocksWith1pOccHU'].iat[0], 2)

# Table 20
row = 144
row += 2
for c in ['White Alone', 'Black Alone', 'AIAN Alone', 'Asian Alone', 'NHOPI Alone', 'Some Other Race Alone', 'Two Or More Races', 'Hispanic', 'Non-Hispanic White']:
    ws3['B'+str(row)] = df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == 'All') & (df['Characteristic'] == f"HDF 20+ Year Median Age Diff {c}"), 'NumCells'].iat[0]
    ws3['C'+str(row)] = df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == 'All') & (df['Characteristic'] == f"HDF 20+ Year Median Age Diff {c}"), 'Inconsistent'].iat[0]
    ws3['D'+str(row)] = round(100* (df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == 'All') & (df['Characteristic'] == f"HDF 20+ Year Median Age Diff {c}"), 'Inconsistent'].iat[0])/df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == 'All') & (df['Characteristic'] == f"HDF 20+ Year Median Age Diff {c}"), 'NumCells'].iat[0], 2)
    ws3['E'+str(row)] = df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == 'All') & (df['Characteristic'] == f"MDF 20+ Year Median Age Diff {c}"), 'NumCells'].iat[0]
    ws3['F'+str(row)] = df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == 'All') & (df['Characteristic'] == f"MDF 20+ Year Median Age Diff {c}"), 'Inconsistent'].iat[0]
    ws3['G'+str(row)] = round(100* (df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == 'All') & (df['Characteristic'] == f"MDF 20+ Year Median Age Diff {c}"), 'Inconsistent'].iat[0])/df.loc[(df['Geography'] == 'County') & (df['Size_Category'] == 'All') & (df['Characteristic'] == f"MDF 20+ Year Median Age Diff {c}"), 'NumCells'].iat[0], 2)
    row += 1
row = 155
row += 2
for c in ['White Alone', 'Black Alone', 'AIAN Alone', 'Asian Alone', 'NHOPI Alone', 'Some Other Race Alone', 'Two Or More Races', 'Hispanic', 'Non-Hispanic White']:
    ws3['B'+str(row)] = df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == f"HDF 20+ Year Median Age Diff {c}"), 'NumCells'].iat[0]
    ws3['C'+str(row)] = df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == f"HDF 20+ Year Median Age Diff {c}"), 'Inconsistent'].iat[0]
    ws3['D'+str(row)] = round(100* (df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == f"HDF 20+ Year Median Age Diff {c}"), 'Inconsistent'].iat[0])/df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == f"HDF 20+ Year Median Age Diff {c}"), 'NumCells'].iat[0], 2)
    ws3['E'+str(row)] = df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == f"MDF 20+ Year Median Age Diff {c}"), 'NumCells'].iat[0]
    ws3['F'+str(row)] = df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == f"MDF 20+ Year Median Age Diff {c}"), 'Inconsistent'].iat[0]
    ws3['G'+str(row)] = round(100* (df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == f"MDF 20+ Year Median Age Diff {c}"), 'Inconsistent'].iat[0])/df.loc[(df['Geography'] == 'Tract') & (df['Size_Category'] == 'All') & (df['Characteristic'] == f"MDF 20+ Year Median Age Diff {c}"), 'NumCells'].iat[0], 2)
    row += 1

if unitmetricsready:
# Table 21
    row = 171
    ws3['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "HDF Under 18 In All Adult GQ Tracts"), 'NumCells'].iat[0]
    ws3['C'+str(row)] = dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "HDF Under 18 In All Adult GQ Tracts"), 'Inconsistent'].iat[0]
    ws3['D'+str(row)] = round(100* (dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "HDF Under 18 In All Adult GQ Tracts"), 'Inconsistent'].iat[0])/dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "HDF Under 18 In All Adult GQ Tracts"), 'NumCells'].iat[0], 2)
    ws3['E'+str(row)] = dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "MDF Under 18 In All Adult GQ Tracts"), 'NumCells'].iat[0]
    ws3['F'+str(row)] = dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "MDF Under 18 In All Adult GQ Tracts"), 'Inconsistent'].iat[0]
    ws3['G'+str(row)] = round(100* (dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "MDF Under 18 In All Adult GQ Tracts"), 'Inconsistent'].iat[0])/dfu.loc[(dfu['Geography'] == 'Tract') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "MDF Under 18 In All Adult GQ Tracts"), 'NumCells'].iat[0], 2)

if unitmetricsready:
    # Table 22
    row = 177
    ws3['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'Block') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumCells'].iat[0]
    ws3['C'+str(row)] = dfu.loc[(dfu['Geography'] == 'Block') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumberHundredOnMDFNotHDF'].iat[0]
    ws3['D'+str(row)] = round(100* (dfu.loc[(dfu['Geography'] == 'Block') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumberHundredOnMDFNotHDF'].iat[0])/dfu.loc[(dfu['Geography'] == 'Block') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumCells'].iat[0], 2)
    # Table 23
    row = 183
    ws3['B'+str(row)] = dfu.loc[(dfu['Geography'] == 'Block') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumCells'].iat[0]
    ws3['C'+str(row)] = dfu.loc[(dfu['Geography'] == 'Block') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumberZeroOnMDFNotHDF'].iat[0]
    ws3['D'+str(row)] = round(100* (dfu.loc[(dfu['Geography'] == 'Block') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumberZeroOnMDFNotHDF'].iat[0])/dfu.loc[(dfu['Geography'] == 'Block') & (dfu['Size_Category'] == 'All') & (dfu['Characteristic'] == "Occupancy Rate"), 'NumCells'].iat[0], 2)



wb.save(filename =  f"{OUTPUTDIR}/metrics{OUTPUTFILEEND}.xlsx")
